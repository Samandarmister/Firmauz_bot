import sqlite3
import re
import os
import openpyxl
from database import get_firma_name, check_firma, get_manual_report, check_file, get_user_language
from config import DATA_PATH
from lang import get_text, get_month_name, translate_text
from converters import convert_to_cyrillic
import logging

logger = logging.getLogger(__name__)

def parse_yagona_excel(file_path, lang='uz_latin'):
    try:
        # Fayl mavjudligini tekshirish
        if not os.path.exists(file_path):
            logger.error(f"Fayl topilmadi: {file_path}")
            return None, f"Fayl topilmadi: {file_path}"

        workbook = openpyxl.load_workbook(file_path)
        # Tilga qarab varag‘ nomini tanlash
        sheet_name = 'Sheet1' if lang == 'uz_latin' else 'Лист1'
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            logger.error(f"Varag‘ topilmadi: {sheet_name}")
            return None, f"Excel faylni o‘qishda xato: Varag‘ '{sheet_name}' topilmadi."

        firms = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            stir, oy, firma_nomi, rahbar, soliq_turi_yagona, yil_boshidan_aylanma, shu_oy_aylanma = row

            if not stir or not oy or not firma_nomi or not rahbar or not soliq_turi_yagona:
                logger.warning(f"Noto'g'ri qator: {row}")
                continue
            if not re.match(r'^\d{9}$', str(stir)):
                logger.warning(f"Noto'g'ri STIR: {stir}")
                continue
            if not check_firma(str(stir)):
                logger.warning(f"STIR ma'lumotlar bazasida yo'q: {stir}")
                continue

            # Oy ni kichik harfga aylantirish va kirill/lotinni qabul qilish
            oy = str(oy).lower()
            month_map = {
                'январ': 'yanvar', 'феврал': 'fevral', 'март': 'mart', 
                'апрел': 'aprel', 'май': 'may', 'июн': 'iyun', 'июл': 'iyul',
                'yanvar': 'yanvar', 'fevral': 'fevral', 'mart': 'mart', 
                'aprel': 'aprel', 'may': 'may', 'iyun': 'iyun', 'iyul': 'iyul'
            }
            oy = month_map.get(oy, oy)
            if oy not in ['yanvar', 'fevral', 'mart', 'aprel', 'may', 'iyun', 'iyul']:
                logger.warning(f"Noto'g'ri oy: {oy}, original: {row[1]}")
                continue

            # Ma'lumotlarni tilga qarab tarjima qilish
            if lang == 'uz_cyrillic':
                firma_nomi = convert_to_cyrillic(firma_nomi)
                rahbar = convert_to_cyrillic(rahbar)
            else:
                firma_nomi = translate_text(firma_nomi, lang)
                rahbar = translate_text(rahbar, lang)

            key = (str(stir), oy)
            firms[key] = {
                'stir': str(stir),
                'oy': oy,
                'firma_nomi': firma_nomi,
                'rahbar': rahbar,
                'soliq_turi_yagona': soliq_turi_yagona,
                'yil_boshidan_aylanma': int(yil_boshidan_aylanma),
                'shu_oy_aylanma': int(shu_oy_aylanma)
            }

        if not firms:
            logger.warning(f"Faylda ma'lumot topilmadi: {file_path}")
            return None, f"Faylda yagona soliq ma'lumotlari topilmadi."

        return firms, None
    except Exception as e:
        logger.error(f"Yagona Excel parsing xatosi: {e}, fayl: {file_path}")
        return None, f"Yagona Excel faylni o‘qishda xato: {str(e)}"

def parse_qqs_excel(file_path, lang='uz_latin'):
    try:
        # Fayl mavjudligini tekshirish
        if not os.path.exists(file_path):
            logger.error(f"Fayl topilmadi: {file_path}")
            return None, f"Fayl topilmadi: {file_path}"

        workbook = openpyxl.load_workbook(file_path)
        # Tilga qarab varag‘ nomini tanlash
        sheet_name = 'Sheet1' if lang == 'uz_latin' else 'Лист1'
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            logger.error(f"Varag‘ topilmadi: {sheet_name}")
            return None, f"Excel faylni o‘qishda xato: Varag‘ '{sheet_name}' topilmadi."

        firms = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            stir, oy, firma_nomi, rahbar, soliq_turi_qqs, yil_boshidan_qqs, shu_oy_qqs = row

            if not stir or not oy or not firma_nomi or not rahbar or not soliq_turi_qqs:
                logger.warning(f"Noto'g'ri qator: {row}")
                continue
            if not re.match(r'^\d{9}$', str(stir)):
                logger.warning(f"Noto'g'ri STIR: {stir}")
                continue
            if not check_firma(str(stir)):
                logger.warning(f"STIR ma'lumotlar bazasida yo'q: {stir}")
                continue

            # Oy ni kichik harfga aylantirish va kirill/lotinni qabul qilish
            oy = str(oy).lower()
            month_map = {
                'январ': 'yanvar', 'феврал': 'fevral', 'март': 'mart', 
                'апрел': 'aprel', 'май': 'may', 'июн': 'iyun', 'июл': 'iyul',
                'yanvar': 'yanvar', 'fevral': 'fevral', 'mart': 'mart', 
                'aprel': 'aprel', 'may': 'may', 'iyun': 'iyun', 'iyul': 'iyul'
            }
            oy = month_map.get(oy, oy)
            if oy not in ['yanvar', 'fevral', 'mart', 'aprel', 'may', 'iyun', 'iyul']:
                logger.warning(f"Noto'g'ri oy: {oy}, original: {row[1]}")
                continue

            # Ma'lumotlarni tilga qarab tarjima qilish
            if lang == 'uz_cyrillic':
                firma_nomi = convert_to_cyrillic(firma_nomi)
                rahbar = convert_to_cyrillic(rahbar)
            else:
                firma_nomi = translate_text(firma_nomi, lang)
                rahbar = translate_text(rahbar, lang)

            key = (str(stir), oy)
            firms[key] = {
                'stir': str(stir),
                'oy': oy,
                'firma_nomi': firma_nomi,
                'rahbar': rahbar,
                'soliq_turi_qqs': soliq_turi_qqs,
                'yil_boshidan_qqs': int(yil_boshidan_qqs),
                'shu_oy_qqs': int(shu_oy_qqs)
            }

        if not firms:
            logger.warning(f"Faylda ma'lumot topilmadi: {file_path}")
            return None, f"Faylda QQS ma'lumotlari topilmadi."

        return firms, None
    except Exception as e:
        logger.error(f"QQS Excel parsing xatosi: {e}, fayl: {file_path}")
        return None, f"QQS Excel faylni o‘qishda xato: {str(e)}"

def generate_yagona_summary(stir, oy, lang='uz_latin'):
    try:
        conn = sqlite3.connect(os.path.join(DATA_PATH, "bot.db"))
        c = conn.cursor()
        c.execute("SELECT name, rahbar, ys_stavka FROM firms WHERE stir = ?", (stir,))
        result = c.fetchone()
        conn.close()

        if not result:
            return translate_text("❌ Firma topilmadi.", lang)

        firma_nomi, rahbar, ys_stavka = result
        if lang == 'uz_cyrillic':
            firma_nomi = convert_to_cyrillic(firma_nomi)
            rahbar = convert_to_cyrillic(rahbar)

        # Tilga qarab fayl nomini tanlash
        file_name = f"{get_month_name(lang, oy)}1.xlsx" if lang == 'uz_latin' else f"Май1.xlsx"
        file_path = os.path.join(DATA_PATH, stir, "yagona", file_name)
        logger.info(f"Yagona fayl yo‘li: {file_path}")

        firms, error = parse_yagona_excel(file_path, lang)
        if error or not firms:
            return translate_text(f"❌ Yagona hisoboti uchun ma'lumot topilmadi: {error or 'Malumotlar topilmadi'}", lang)

        key = (stir, oy.lower())
        if key not in firms:
            return translate_text(f"❌ {get_month_name(lang, oy)} uchun yagona hisoboti topilmadi.", lang)

        firm = firms[key]
        yil_boshidan_aylanma = firm['yil_boshidan_aylanma']
        shu_oy_aylanma = firm['shu_oy_aylanma']
        soliq_turi_yagona = firm['soliq_turi_yagona']
        yagona_soliq = int(shu_oy_aylanma * (float(soliq_turi_yagona.strip('%')) / 100))

        return get_text(
            lang,
            'yagona_report',
            firma_nomi=firma_nomi,
            rahbar=rahbar,
            oy=get_month_name(lang, oy),
            yil_boshidan_aylanma=f"{yil_boshidan_aylanma:,}",
            shu_oy_aylanma=f"{shu_oy_aylanma:,}",
            soliq_turi_yagona=soliq_turi_yagona,
            yagona_soliq=f"{yagona_soliq:,}"
        )
    except Exception as e:
        logger.error(f"Yagona hisoboti yaratishda xato: {e}, STIR={stir}, Oy={oy}")
        return translate_text(f"❌ Hisobot yaratishda xato: {str(e)}", lang)

def generate_qqs_summary(stir, oy, lang='uz_latin'):
    try:
        conn = sqlite3.connect(os.path.join(DATA_PATH, "bot.db"))
        c = conn.cursor()
        c.execute("SELECT name, rahbar, qqs_stavka FROM firms WHERE stir = ?", (stir,))
        result = c.fetchone()
        conn.close()

        if not result:
            return translate_text("❌ Firma topilmadi.", lang)

        firma_nomi, rahbar, qqs_stavka = result
        if lang == 'uz_cyrillic':
            firma_nomi = convert_to_cyrillic(firma_nomi)
            rahbar = convert_to_cyrillic(rahbar)

        # Tilga qarab fayl nomini tanlash
        file_name = f"{get_month_name(lang, oy)}1.xlsx" if lang == 'uz_latin' else f"Май1.xlsx"
        file_path = os.path.join(DATA_PATH, stir, "qqs", file_name)
        logger.info(f"QQS fayl yo‘li: {file_path}")

        firms, error = parse_qqs_excel(file_path, lang)
        if error or not firms:
            return translate_text(f"❌ QQS hisoboti uchun ma'lumot topilmadi: {error or 'Malumotlar topilmadi'}", lang)

        key = (stir, oy.lower())
        if key not in firms:
            return translate_text(f"❌ {get_month_name(lang, oy)} uchun QQS hisoboti topilmadi.", lang)

        firm = firms[key]
        yil_boshidan_qqs = firm['yil_boshidan_qqs']
        shu_oy_qqs = firm['shu_oy_qqs']
        soliq_turi_qqs = firm['soliq_turi_qqs']
        qqs_soliq = int(shu_oy_qqs * (float(soliq_turi_qqs.strip('%')) / 100))

        return get_text(
            lang,
            'qqs_report',
            firma_nomi=firma_nomi,
            rahbar=rahbar,
            oy=get_month_name(lang, oy),
            yil_boshidan_qqs=f"{yil_boshidan_qqs:,}",
            shu_oy_qqs=f"{shu_oy_qqs:,}",
            soliq_turi_qqs=soliq_turi_qqs,
            qqs_soliq=f"{qqs_soliq:,}"
        )
    except Exception as e:
        logger.error(f"QQS hisoboti yaratishda xato: {e}, STIR={stir}, Oy={oy}")
        return translate_text(f"❌ Hisobot yaratishda xato: {str(e)}", lang)
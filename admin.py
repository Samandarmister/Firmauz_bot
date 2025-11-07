import os
import openpyxl
import re
from lang import translate_text
import logging
from openpyxl import load_workbook
import sqlite3
from datetime import datetime
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from loader import dp, bot
from config import ADMIN_IDS, DATA_PATH
from database import (
    add_firma, check_firma, get_all_firms, save_file, check_file,
    save_manual_report, get_firma_name, get_user_language, get_firma_info,
    save_qqs_report, save_yagona_report, verify_owner_phone, get_firm_docs,
    today_downloads, log_alert, save_firm_docs, log_download,
    add_firm_owner
)
from lang import get_text, get_month_name, translate_text
from converters import convert_to_cyrillic, convert_to_latin

logging.basicConfig(level=logging.INFO, filename="bot.log", encoding="utf-8")
logger = logging.getLogger(__name__)

class VerifyPhone(StatesGroup):
    stir = State()
    phone = State()


class UploadFirmDocs(StatesGroup):
    stir = State()
    pdf1 = State()
    pdf2 = State()
    pfx = State()

class AddFirma(StatesGroup):
    stir = State()
    name = State()
    soliq_turi = State()
    ds_stavka = State()
    ys_stavka = State()
    qqs_stavka = State()
    rahbar = State()   # ✅ Yangi
    phone = State()    # ✅ Yangi
    
class EditFirma(StatesGroup):
    stir = State()
    new_name = State()

class UploadFiles(StatesGroup):
    soliq_turi = State()
    oy = State()
    excel1 = State()
    excel2 = State()
    html = State()

class DeleteReport(StatesGroup):
    stir = State()
    oy = State()

class ManualInput(StatesGroup):
    select_soliq_turi = State()  # Soliq turini tanlash
    excel_upload = State()       # Excel fayl yuklash
    stir = State()               # Firma STIR raqami
    oy = State()                 # Oy tanlash
    firma_name = State()         # Firma nomi
    xodimlar_soni = State()      # Xodimlar soni (daromad uchun)
    xodimlar_data = State()      # Xodimlar ma'lumotlari (daromad uchun)
    yagona_data = State()        # Yagona soliq ma'lumotlari
    qqs_data = State()           # QQS ma'lumotlari
    confirm = State()            # Tasdiqlash
    search = State()             # Qidiruv

def is_admin(user_id):
    return user_id in ADMIN_IDS

def parse_excel_file(file_path, lang='uz_latin'):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        firms = {}
        current_stir = None
        current_oy = None
        current_firma_nomi = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            stir = row[0] if row[0] else current_stir
            oy = row[1] if row[1] else current_oy
            firma_nomi = row[2] if row[2] else current_firma_nomi
            lavozim = row[3]
            ism = row[4]
            yil_boshidan = row[5]
            shu_oy = row[6]

            if not stir or not oy or not firma_nomi or not lavozim or not ism or yil_boshidan is None or shu_oy is None:
                logger.warning(f"Noto'g'ri qator: {row}")
                continue

            if not isinstance(yil_boshidan, (int, float)) or not isinstance(shu_oy, (int, float)):
                logger.warning(f"Noto'g'ri maosh formati: {row}")
                continue

            oy = oy.lower()
            if oy not in ['yanvar', 'fevral', 'mart', 'aprel', 'may', 'iyun', 'iyul']:
                logger.warning(f"Noto'g'ri oy: {oy}")
                continue

            if not re.match(r'^\d{9}$', str(stir)):
                logger.warning(f"Noto'g'ri STIR: {stir}")
                continue

            if not check_firma(str(stir)):
                logger.warning(f"STIR ma'lumotlar bazasida yo'q: {stir}")
                continue

            current_stir = str(stir)
            current_oy = oy
            if lang == 'uz_cyrillic':
                current_firma_nomi = convert_to_cyrillic(firma_nomi)
                lavozim = convert_to_cyrillic(lavozim)
                ism = convert_to_cyrillic(ism)
            else:
                current_firma_nomi = translate_text(firma_nomi, lang)
                lavozim = translate_text(lavozim, lang)
                ism = translate_text(ism, lang)

            key = (current_stir, current_oy)
            if key not in firms:
                firms[key] = {
                    'stir': current_stir,
                    'oy': current_oy,
                    'firma_nomi': current_firma_nomi,
                    'xodimlar': []
                }

            firms[key]['xodimlar'].append({
                'lavozim': lavozim,
                'ism': ism,
                'yil_boshidan': int(yil_boshidan),
                'shu_oy': int(shu_oy)
            })

        return firms, None
    except Exception as e:
        logger.error(f"Excel parsing xatosi: {e}")
        return None, f"Excel faylni o'qishda xato: {str(e)}"
    


def generate_firma_excel(stir, oy, firma_nomi, xodimlar, dest_path_latin, dest_path_cyrillic):
    try:
        # Lotin tilida fayl yaratish
        workbook_latin = openpyxl.Workbook()
        sheet_latin = workbook_latin.active
        sheet_latin.title = "Sheet1"
        headers_latin = [
            translate_text("STIR", 'uz_latin'),
            translate_text("Oy", 'uz_latin'),
            translate_text("Firma nomi", 'uz_latin'),
            translate_text("Xodim lavozimi", 'uz_latin'),
            translate_text("Ism Familyasi", 'uz_latin'),
            translate_text("Yil boshidan", 'uz_latin'),
            translate_text("Shu Oy uchun oylik", 'uz_latin')
        ]
        sheet_latin.append(headers_latin)

        for i, xodim in enumerate(xodimlar):
            row = [
                stir if i == 0 else "",
                get_month_name('uz_latin', oy) if i == 0 else "",
                translate_text(firma_nomi, 'uz_latin') if i == 0 else "",
                translate_text(xodim['lavozim'], 'uz_latin'),
                translate_text(xodim['ism'], 'uz_latin'),
                xodim['yil_boshidan'],
                xodim['shu_oy']
            ]
            sheet_latin.append(row)

        os.makedirs(os.path.dirname(dest_path_latin), exist_ok=True)
        workbook_latin.save(dest_path_latin)
        logger.info(f"Yangi Excel fayli yaratildi (lotin): {dest_path_latin}")

        # Kirill tilida fayl yaratish
        workbook_cyrillic = openpyxl.Workbook()
        sheet_cyrillic = workbook_cyrillic.active
        sheet_cyrillic.title = "Лист1"
        headers_cyrillic = [
            translate_text("STIR", 'uz_cyrillic'),
            translate_text("Oy", 'uz_cyrillic'),
            translate_text("Firma nomi", 'uz_cyrillic'),
            translate_text("Xodim lavozimi", 'uz_cyrillic'),
            translate_text("Ism Familyasi", 'uz_cyrillic'),
            translate_text("Yil boshidan", 'uz_cyrillic'),
            translate_text("Shu Oy uchun oylik", 'uz_cyrillic')
        ]
        sheet_cyrillic.append(headers_cyrillic)

        for i, xodim in enumerate(xodimlar):
            row = [
                stir if i == 0 else "",
                get_month_name('uz_cyrillic', oy) if i == 0 else "",
                translate_text(firma_nomi, 'uz_cyrillic') if i == 0 else "",
                translate_text(xodim['lavozim'], 'uz_cyrillic'),
                translate_text(xodim['ism'], 'uz_cyrillic'),
                xodim['yil_boshidan'],
                xodim['shu_oy']
            ]
            sheet_cyrillic.append(row)

        os.makedirs(os.path.dirname(dest_path_cyrillic), exist_ok=True)
        workbook_cyrillic.save(dest_path_cyrillic)
        logger.info(f"Yangi Excel fayli yaratildi (kirill): {dest_path_cyrillic}")

        return True
    except Exception as e:
        logger.error(f"Excel faylini yaratishda xato: {e}")
        return False

def create_paginated_keyboard(items, callback_prefix, page=1, per_page=10, lang='uz_latin'):
    total_items = len(items)
    total_pages = (total_items + per_page - 1) // per_page
    page = max(1, min(page, total_pages))
    start_idx = (page - 1) * per_page
    end_idx = min(start_idx + per_page, total_items)

    keyboard = InlineKeyboardMarkup(row_width=2)
    for item in items[start_idx:end_idx]:
        if isinstance(item, tuple) and len(item) == 2:
            stir, name = item
            keyboard.add(InlineKeyboardButton(f"{translate_text(name, lang)} ({stir})", callback_data=f"{callback_prefix}_{stir}"))
        elif isinstance(item, tuple) and len(item) == 3:
            stir, oy, firma_nomi = item
            keyboard.add(InlineKeyboardButton(f"{translate_text(firma_nomi, lang)} ({stir}, {get_month_name(lang, oy)})", callback_data=f"{callback_prefix}_{stir}_{oy}"))

    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton(translate_text("⬅️ Oldingi", lang), callback_data=f"{callback_prefix}_page_{page-1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton(translate_text("Keyingi ➡️", lang), callback_data=f"{callback_prefix}_page_{page+1}"))
    if nav_buttons:
        keyboard.row(*nav_buttons)

    keyboard.add(InlineKeyboardButton(translate_text("🔍 Qidirish", lang), callback_data=f"{callback_prefix}_search"))
    keyboard.add(InlineKeyboardButton(translate_text("🔙 Orqaga", lang), callback_data="back_to_admin"))

    return keyboard, page, total_pages


def back_to_admin_keyboard(lang):
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton(translate_text("Admin paneliga qaytish", lang), callback_data="back_to_admin")
    )
    return keyboard


@dp.callback_query_handler(lambda c: c.data.startswith("list_firmas_page_"), user_id=ADMIN_IDS)
async def list_firmas(callback_query: types.CallbackQuery):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    page = int(callback_query.data.split("_")[-1])
    per_page = 10

    # Ma'lumotlar bazasidan firmalarni olish
    conn = sqlite3.connect(os.path.join(DATA_PATH, "bot.db"))
    c = conn.cursor()
    c.execute("SELECT stir, name FROM firms ORDER BY name LIMIT ? OFFSET ?", (per_page, (page - 1) * per_page))
    firms = c.fetchall()
    c.execute("SELECT COUNT(*) FROM firms")
    total_firms = c.fetchone()[0]
    conn.close()

    if not firms:
        await callback_query.message.edit_text(
            translate_text("❌ Hozirda hech qanday firma mavjud emas.", lang),
            reply_markup=back_to_admin_keyboard(lang)
        )
        return

    # ✅ Shu yerda firms_list ni tuzish kerak
    firms_list = [f"{i + 1 + (page - 1) * per_page}. {name} (STIR: {stir})" for i, (stir, name) in enumerate(firms)]
    firmalar_text = "\n".join(firms_list)
    total_pages = (total_firms + per_page - 1) // per_page

    response = (
        f"📋 {translate_text('Firmalar ro‘yxati', lang)} ({total_firms} ta):\n\n"
        f"{firmalar_text}\n\n"
        f"📄 {translate_text('Sahifa', lang)}: {page}/{total_pages}"
    )

    # Navigatsiya tugmalari
    keyboard = InlineKeyboardMarkup(row_width=3)
    if page > 1:
        keyboard.insert(InlineKeyboardButton("⬅️", callback_data=f"list_firmas_page_{page - 1}"))
    if page < total_pages:
        keyboard.insert(InlineKeyboardButton("➡️", callback_data=f"list_firmas_page_{page + 1}"))
    keyboard.add(InlineKeyboardButton(translate_text("Admin paneliga qaytish", lang), callback_data="back_to_admin"))

    await callback_query.message.edit_text(response, reply_markup=keyboard)


@dp.callback_query_handler(lambda c: c.data == "back_to_admin", user_id=ADMIN_IDS)
async def back_to_admin_handler(callback_query: types.CallbackQuery, state: FSMContext):
    await back_to_admin_panel(callback_query, state)

@dp.message_handler(commands=['admin'], user_id=ADMIN_IDS)
async def admin_panel(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    logger.info(f"Admin panel: user_id={user_id}, lang={lang}")
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Yangi firma qo'shish", lang), callback_data="add_firma"),
        InlineKeyboardButton(translate_text("Excel orqali firmalar qo'shish", lang), callback_data="add_firms_excel"),
        InlineKeyboardButton(translate_text("Firma tahrirlash", lang), callback_data="edit_firma"),
        InlineKeyboardButton(translate_text("Fayl yuklash", lang), callback_data="upload_files"),
        InlineKeyboardButton(translate_text("Qo'lda hisobot kiritish", lang), callback_data="manual_input"),
        InlineKeyboardButton(translate_text("Hisobot o'chirish", lang), callback_data="delete_report"),
        InlineKeyboardButton(translate_text("Firmalar ro'yxati", lang), callback_data="list_firmas_page_1"),
        InlineKeyboardButton(translate_text("📎 Firma hujjat yuklash", lang), callback_data="upload_firm_docs"),
        InlineKeyboardButton("📞 Telefon o‘zgartirish", callback_data="edit_firm_phone")        
    )
    sent_message = await message.answer(translate_text("🔒 Admin paneliga xush kelibsiz!", lang), reply_markup=keyboard)
    await state.update_data(last_message_id=sent_message.message_id)
    await message.delete()

@dp.callback_query_handler(lambda c: c.data == "edit_firm_phone", user_id=ADMIN_IDS)
async def edit_firm_phone_start(call: types.CallbackQuery, state: FSMContext):
    lang = get_user_language(call.from_user.id)
    await call.message.answer(translate_text("✍️ Firma STIR raqamini kiriting:", lang))
    await EditFirmPhone.waiting_for_stir.set()

class EditFirmPhone(StatesGroup):
    waiting_for_stir = State()
    waiting_for_new_phone = State()

@dp.message_handler(state=EditFirmPhone.waiting_for_stir, user_id=ADMIN_IDS)
async def ask_new_phone(message: types.Message, state: FSMContext):
    stir = message.text.strip()

    if not stir.isdigit() or len(stir) != 9:
        await message.answer("❌ STIR 9 ta raqam bo‘lishi kerak. Qayta kiriting.")
        return

    firma = get_firma_info(stir)
    if not firma:
        await message.answer("❌ Bu STIR bo‘yicha firma topilmadi.")
        await state.finish()
        return

    await state.update_data(stir=stir)
    await message.answer("📱 Yangi telefon raqamini kiriting (masalan: +998901234567):")
    await EditFirmPhone.waiting_for_new_phone.set()



@dp.message_handler(state=AddFirma.qqs_stavka, user_id=ADMIN_IDS)
async def add_firma_qqs(message: types.Message, state: FSMContext):
    await state.update_data(qqs_stavka=message.text)
    await AddFirma.rahbar.set()
    await message.answer("Rahbar F.I.SH ni kiriting:")



@dp.message_handler(state=AddFirma.rahbar, user_id=ADMIN_IDS)
async def add_firma_rahbar(message: types.Message, state: FSMContext):
    await state.update_data(rahbar=message.text)
    await AddFirma.phone.set()
    await message.answer("Rahbar telefon raqamini kiriting (masalan: +998901234567):")


def add_firm_owner(stir, phone):
    conn = sqlite3.connect(os.path.join(DATA_PATH, "bot.db"))
    c = conn.cursor()
    c.execute("SELECT 1 FROM firm_owners WHERE phone=?", (phone,))
    if not c.fetchone():
        c.execute("INSERT INTO firm_owners (stir, phone) VALUES (?, ?)", (stir, phone))
        conn.commit()
    conn.close()






@dp.message_handler(state=VerifyPhone.phone)
async def check_phone(message: types.Message, state: FSMContext):
    phone = message.text.strip()
    data = await state.get_data()
    stir = data['stir']

    if not verify_owner_phone(stir, phone):
        log_alert(message.from_user.id, phone, stir, "Unauthorized attempt")
        await state.finish()
        return await message.answer("❌ Sizda ruxsat yo'q.")

    if today_downloads(phone, stir) >= 3:
        return await message.answer("⚠️ Bugun fayllarni 3 martadan ko‘p yuklab bo‘lmaysiz.")

    files = get_firm_docs(stir)
    if not files:
        return await message.answer("❌ Firma hujjatlari hali yuklanmagan.")

    pdf1, pdf2, pfx = files

    for fpath in [pdf1, pdf2, pfx]:
        if fpath and os.path.exists(fpath):
            await message.answer_document(open(fpath, 'rb'))
            log_download(message.from_user.id, phone, stir, fpath)

    await state.finish()
    await message.answer("✅ Hujjatlar yuborildi.")

@dp.callback_query_handler(lambda c: c.data=="upload_firm_docs", user_id=ADMIN_IDS)
async def start_upload_docs(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    await UploadFirmDocs.stir.set()
    await callback.message.answer("Firma STIR raqamini kiriting:")

@dp.message_handler(state=UploadFirmDocs.stir, user_id=ADMIN_IDS)
async def docs_stir(message: types.Message, state: FSMContext):
    stir = message.text.strip()
    if not check_firma(stir):
        await message.answer("❌ Bunday STIR bazada yo‘q!")
        return
    await state.update_data(stir=stir)
    await UploadFirmDocs.pdf1.set()
    await message.answer("1-PDF faylni yuboring:")



@dp.message_handler(content_types=['document'], state=UploadFirmDocs.pdf1, user_id=ADMIN_IDS)
async def docs_pdf1(message: types.Message, state: FSMContext):
    data = await state.get_data()
    stir = data["stir"]
    
    path = f"{DATA_PATH}/{stir}/firm_docs"
    os.makedirs(path, exist_ok=True)

    save_to = f"{path}/doc1.pdf"
    await message.document.download(destination_file=save_to)

    await state.update_data(pdf1=save_to)
    await UploadFirmDocs.pdf2.set()
    await message.answer("Ikkinchi PDF faylni yuboring:")

@dp.message_handler(content_types=['document'], state=UploadFirmDocs.pdf2, user_id=ADMIN_IDS)
async def docs_pdf2(message: types.Message, state: FSMContext):
    data = await state.get_data()
    stir = data["stir"]

    save_to = f"{DATA_PATH}/{stir}/firm_docs/doc2.pdf"
    await message.document.download(destination_file=save_to)

    await state.update_data(pdf2=save_to)
    await UploadFirmDocs.pfx.set()
    await message.answer("PFX faylni yuboring:")

@dp.message_handler(content_types=['document'], state=UploadFirmDocs.pfx, user_id=ADMIN_IDS)
async def docs_pfx(message: types.Message, state: FSMContext):
    data = await state.get_data()
    stir = data["stir"]

    path = f"{DATA_PATH}/{stir}/firm_docs"
    os.makedirs(path, exist_ok=True)

    # asl fayl nomini olamiz
    original_name = message.document.file_name
    save_to = f"{path}/{original_name}"

    await message.document.download(destination_file=save_to)

    save_firm_docs(stir, data["pdf1"], data["pdf2"], save_to)

    await state.finish()
    await message.answer("✅ Firma hujjatlari yuklandi va saqlandi!")




@dp.message_handler(lambda m: m.text.isdigit() and len(m.text)==9)
async def user_send_stir(message: types.Message, state: FSMContext):
    stir = message.text

    if not check_firma(stir):
        return await message.answer("❌ Bunday firma topilmadi.")

    await state.update_data(stir=stir)
    await VerifyPhone.phone.set()
    await message.answer("Telefon raqamingizni +998XXXXXXXXX formatda kiriting:")


@dp.callback_query_handler(lambda c: c.data == "back_to_admin", user_id=ADMIN_IDS)
async def back_to_admin_panel(callback_query: types.CallbackQuery = None, state: FSMContext = None):
    if callback_query:
        await bot.answer_callback_query(callback_query.id)
        user_id = callback_query.from_user.id
        message = callback_query.message
    else:
        if state:
            data = await state.get_data()
            user_id = data.get('user_id')
            message = None
        else:
            return

    lang = get_user_language(user_id)
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Yangi firma qo'shish", lang), callback_data="add_firma"),
        InlineKeyboardButton(translate_text("Excel orqali firmalar qo'shish", lang), callback_data="add_firms_excel"),
        InlineKeyboardButton(translate_text("Firma tahrirlash", lang), callback_data="edit_firma"),
        InlineKeyboardButton(translate_text("Fayl yuklash", lang), callback_data="upload_files"),
        InlineKeyboardButton(translate_text("Qo'lda hisobot kiritish", lang), callback_data="manual_input"),
        InlineKeyboardButton(translate_text("Hisobot o'chirish", lang), callback_data="delete_report"),
        InlineKeyboardButton(translate_text("Firmalar ro'yxati", lang), callback_data="list_firmas_page_1")
    )

    try:
        if message:
            await message.delete()  # Avvalgi xabarni o‘chirish
        sent_message = await bot.send_message(user_id, translate_text("🔒 Admin paneliga xush kelibsiz!", lang), reply_markup=keyboard)
        await state.update_data(last_message_id=sent_message.message_id)
    except Exception as e:
        logger.error(f"Xabar o'chirish/yuborishda xato: {e}")
        await bot.send_message(user_id, translate_text("❌ Xatolik yuz berdi, qayta urinib ko'ring.", lang))

    if state:
        await state.finish()


async def some_callback_handler(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    last_message_id = data.get('last_message_id')

    # Avvalgi xabarni o‘chirish
    if last_message_id:
        try:
            await bot.delete_message(chat_id=user_id, message_id=last_message_id)
        except Exception as e:
            logger.warning(f"Xabar o'chirishda xato: {e}, message_id={last_message_id}")

    # InlineKeyboardMarkup obyektini yaratish
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Orqaga", lang), callback_data="back_to_admin")
    )

    # Yangi xabar yuborish va ID sini saqlash
    sent_message = await bot.send_message(
        user_id,
        translate_text("Xabar matni", lang),
        reply_markup=keyboard
    )
    await state.update_data(last_message_id=sent_message.message_id)
    # Qolgan logika...


@dp.callback_query_handler(lambda c: c.data == "add_firma", user_id=ADMIN_IDS)
async def start_add_firma(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    last_message_id = data.get('last_message_id')

    # Avvalgi xabarni o‘chirish
    if last_message_id:
        try:
            await bot.delete_message(chat_id=user_id, message_id=last_message_id)
        except Exception as e:
            logger.warning(f"Xabar o'chirishda xato: {e}, message_id={last_message_id}")

    await AddFirma.stir.set()
    sent_message = await bot.send_message(user_id, translate_text("Yangi firma STIR raqamini kiriting (9 raqam, masalan: 302824863):", lang))
    await state.update_data(last_message_id=sent_message.message_id)


@dp.message_handler(commands=['cancel'], user_id=ADMIN_IDS, state='*')
async def cancel_operation(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    last_message_id = data.get('last_message_id')
    excel_file_path = data.get('excel_file_path')

    # Avvalgi inline xabarni o‘chirish
    if last_message_id:
        try:
            await bot.delete_message(chat_id=user_id, message_id=last_message_id)
        except Exception as e:
            logger.warning(f"Xabar o'chirishda xato: {e}, message_id={last_message_id}")

    # Vaqtinchalik faylni o‘chirish
    if excel_file_path and os.path.exists(excel_file_path):
        try:
            os.remove(excel_file_path)
            logger.info(f"Vaqtinchalik fayl o'chirildi: {excel_file_path}")
        except Exception as e:
            logger.error(f"Vaqtinchalik fayl o'chirishda xato: {e}, path={excel_file_path}")

    # Holatni tozalash
    await state.finish()

    # Admin panelini qayta chiqarish
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Yangi firma qo'shish", lang), callback_data="add_firma"),
        InlineKeyboardButton(translate_text("Excel orqali firmalar qo'shish", lang), callback_data="add_firms_excel"),
        InlineKeyboardButton(translate_text("Firma tahrirlash", lang), callback_data="edit_firma"),
        InlineKeyboardButton(translate_text("Fayl yuklash", lang), callback_data="upload_files"),
        InlineKeyboardButton(translate_text("Qo'lda hisobot kiritish", lang), callback_data="manual_input"),
        InlineKeyboardButton(translate_text("Hisobot o'chirish", lang), callback_data="delete_report"),
        InlineKeyboardButton(translate_text("Firmalar ro'yxati", lang), callback_data="list_firmas_page_1")
    )
    sent_message = await message.answer(translate_text("✅ Amaliyot bekor qilindi, admin paneldasiz.", lang), reply_markup=keyboard)
    await state.update_data(last_message_id=sent_message.message_id)
    await message.delete()
    logger.info(f"Cancel operation: user_id={user_id}")





@dp.message_handler(state=AddFirma.soliq_turi, user_id=ADMIN_IDS)
async def process_soliq_turi(message: types.Message, state: FSMContext):
    soliq_turi = message.text.strip().lower()
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    if soliq_turi not in ['ds-ys', 'ds-qqs']:
        await message.answer(translate_text("❌ Soliq turi 'ds-ys' yoki 'ds-qqs' bo'lishi kerak.", lang))
        return
    await state.update_data(soliq_turi=soliq_turi)
    await AddFirma.name.set()
    await message.answer(translate_text("Firma nomini kiriting (kamida 3 belgi):", lang))

@dp.message_handler(state=AddFirma.stir, user_id=ADMIN_IDS)
async def process_stir(message: types.Message, state: FSMContext):
    stir = message.text.strip()
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    if not re.match(r'^\d{9}$', stir):
        await message.answer(translate_text("❌ STIR 9 raqamdan iborat bo'lishi kerak.", lang))
        return
    if check_firma(stir):
        await message.answer(translate_text("❌ Bu STIR allaqachon mavjud.", lang))
        return
    await state.update_data(stir=stir)
    await AddFirma.soliq_turi.set()
    await message.answer(translate_text("Soliq turini kiriting (ds-ys, ds-qqs):", lang))

# 📌 Firma nomini qabul qilish
@dp.message_handler(state=AddFirma.name, user_id=ADMIN_IDS)
async def process_name(message: types.Message, state: FSMContext):
    name = message.text.strip()
    if len(name) < 3:
        return await message.answer("❌ Firma nomi kamida 3 harf bo'lishi kerak")

    await state.update_data(name=name)
    await message.answer("👤 Rahbar F.I.Sh kiriting:")
    await AddFirma.rahbar.set()



# 📌 Rahbar ism qabul qilish
@dp.message_handler(state=AddFirma.rahbar, user_id=ADMIN_IDS)
async def process_rahbar(message: types.Message, state: FSMContext):
    rahbar = message.text.strip()
    await state.update_data(rahbar=rahbar)

    await message.answer("📞 Rahbar telefon raqamini kiriting (masalan: +998901234567):")
    await AddFirma.phone.set()





@dp.callback_query_handler(lambda c: c.data == "edit_firma", user_id=ADMIN_IDS)
async def start_edit_firma(callback_query: types.CallbackQuery):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    firms = get_all_firms()
    if not firms:
        await bot.send_message(callback_query.from_user.id, translate_text("❌ Hozircha firmalar mavjud emas.", lang))
        return
    keyboard, page, total_pages = create_paginated_keyboard(firms, "edit_firm", page=1, lang=lang)
    await bot.send_message(callback_query.from_user.id, translate_text(f"Tahrir qilmoqchi bo'lgan firmani tanlang (Sahifa {page}/{total_pages}):", lang), reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("edit_firm_page_"), user_id=ADMIN_IDS)
async def edit_firma_paginate(callback_query: types.CallbackQuery):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    page = int(callback_query.data.split("_")[-1])
    firms = get_all_firms()
    keyboard, page, total_pages = create_paginated_keyboard(firms, "edit_firm", page=page, lang=lang)
    await bot.edit_message_text(
        translate_text(f"Tahrir qilmoqchi bo'lgan firmani tanlang (Sahifa {page}/{total_pages}):", lang),
        callback_query.from_user.id,
        callback_query.message.message_id,
        reply_markup=keyboard
    )



@dp.callback_query_handler(lambda c: c.data == "edit_firm_search", user_id=ADMIN_IDS, state="*")
async def start_edit_firma_search(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    await state.finish()
    await ManualInput.search.set()
    await state.update_data(search_context="edit_firma")
    await bot.send_message(callback_query.from_user.id, translate_text("Firma STIR yoki nomini kiriting (qisman moslik uchun):", lang))

@dp.callback_query_handler(lambda c: c.data.startswith("edit_firm_"), user_id=ADMIN_IDS)
async def select_firma_to_edit(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    stir = callback_query.data.split("_", 2)[2]
    await state.update_data(stir=stir)
    firma_name = get_firma_name(stir)
    await EditFirma.new_name.set()
    await bot.send_message(callback_query.from_user.id, translate_text(f"Hozirgi firma nomi: {firma_name}\nYangi nomni kiriting (kamida 3 belgi):", lang))

@dp.message_handler(state=EditFirma.new_name, user_id=ADMIN_IDS)
async def process_new_name(message: types.Message, state: FSMContext):
    new_name = message.text.strip()
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    if len(new_name) < 3:
        await message.answer(translate_text("❌ Firma nomi kamida 3 ta belgidan iborat bo'lishi kerak.", lang))
        return
    data = await state.get_data()
    stir = data['stir']
    conn = sqlite3.connect(os.path.join(DATA_PATH, "bot.db"))
    c = conn.cursor()
    c.execute("UPDATE firms SET name = ? WHERE stir = ?", (new_name, stir))
    conn.commit()
    conn.close()
    await state.finish()
    await message.answer(translate_text(f"✅ Firma nomi o'zgartirildi: {new_name} ({stir})", lang))
    logger.info(f"Firma nomi o'zgartirildi: STIR={stir}, Yangi nom={new_name}")

@dp.callback_query_handler(lambda c: c.data == "upload_files", user_id=ADMIN_IDS)
async def start_upload_files(callback_query: types.CallbackQuery):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    firms = get_all_firms()  # await olib tashlandi
    if not firms:
        await bot.send_message(callback_query.from_user.id, translate_text("❌ Hozircha firmalar mavjud emas.", lang))
        return
    keyboard, page, total_pages = create_paginated_keyboard(firms, "firm_upload", page=1, lang=lang)
    await bot.send_message(callback_query.from_user.id, translate_text(f"Fayl yuklash uchun firma tanlang (Sahifa {page}/{total_pages}):", lang), reply_markup=keyboard)


@dp.callback_query_handler(lambda c: c.data.startswith("firm_upload_page_"), user_id=ADMIN_IDS)
async def upload_files_paginate(callback_query: types.CallbackQuery):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    page = int(callback_query.data.split("_")[-1])
    firms = get_all_firms()  # await olib tashlandi
    keyboard, page, total_pages = create_paginated_keyboard(firms, "firm_upload", page=page, lang=lang)
    await bot.edit_message_text(
        translate_text(f"Fayl yuklash uchun firma tanlang (Sahifa {page}/{total_pages}):", lang),
        callback_query.from_user.id,
        callback_query.message.message_id,
        reply_markup=keyboard
    )

@dp.callback_query_handler(lambda c: c.data == "firm_upload_search", user_id=ADMIN_IDS, state="*")
async def start_upload_files_search(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    await state.finish()
    await ManualInput.search.set()
    await state.update_data(search_context="upload_files")
    await bot.send_message(callback_query.from_user.id, translate_text("Firma STIR yoki nomini kiriting (qisman moslik uchun):", lang))

logger = logging.getLogger(__name__)

@dp.callback_query_handler(lambda c: c.data.startswith("firm_upload_"), user_id=ADMIN_IDS)
async def select_soliq_turi(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    stir = callback_query.data.split("_", 2)[2]
    await state.update_data(stir=stir)
    logger.info(f"select_soliq_turi boshlandi: user_id={user_id}, stir={stir}, lang={lang}")

    # Firma ma'lumotlarini olish
    try:
        conn = sqlite3.connect(os.path.join(DATA_PATH, "bot.db"))
        c = conn.cursor()
        c.execute("SELECT name, rahbar, soliq_turi, ds_stavka, ys_stavka, qqs_stavka FROM firms WHERE stir = ?", (stir,))
        result = c.fetchone()
        conn.close()
    except Exception as e:
        logger.error(f"Ma'lumotlar bazasidan xato: {e}, STIR={stir}")
        await bot.send_message(user_id, translate_text("❌ Ma'lumotlar bazasida xato yuz berdi.", lang), parse_mode='Markdown')
        return

    if not result:
        logger.error(f"Firma topilmadi: STIR={stir}")
        await bot.send_message(user_id, translate_text("❌ Firma topilmadi.", lang), parse_mode='Markdown')
        return

    name, rahbar, soliq_turi, ds_stavka, ys_stavka, qqs_stavka = result
    soliq_turi = soliq_turi.lower() if soliq_turi else 'ds-ys'  # Standart qiymat
    logger.info(f"Firma ma'lumotlari: STIR={stir}, Name={name}, Soliq_turi={soliq_turi}")

    # Firma ma'lumotlarini ko'rsatish
    message_text = get_text(lang, 'firma_info',
                            stir=stir,
                            firma_nomi=name,
                            rahbar=rahbar if rahbar else translate_text("Noma'lum", lang),
                            soliq_turi=soliq_turi,
                            ds_stavka=ds_stavka if ds_stavka else "Noma'lum",
                            ys_stavka=ys_stavka if ys_stavka else "Noma'lum",
                            qqs_stavka=qqs_stavka if qqs_stavka else "Noma'lum") + "\n\n" + \
                   f"📊 STIR: {stir}\n" + \
                   translate_text("Soliq turini tanlang", lang) + ":"

    # Soliq turiga qarab tugmalar
    keyboard = InlineKeyboardMarkup(row_width=2)
    if soliq_turi == 'ds-ys':
        keyboard.add(
            InlineKeyboardButton(text=translate_text("Daromad solig'i", lang), callback_data="upload_daromad"),
            InlineKeyboardButton(text=translate_text("Yagona soliq", lang), callback_data="upload_yagona")
        )
    elif soliq_turi == 'ds-qqs':
        keyboard.add(
            InlineKeyboardButton(text=translate_text("Daromad solig'i", lang), callback_data="upload_daromad"),
            InlineKeyboardButton(text=translate_text("Qo‘shilgan qiymat solig‘i", lang), callback_data="upload_qqs")
        )
    else:
        logger.error(f"Noto'g'ri soliq_turi: {soliq_turi} firma uchun {stir}")
        keyboard.add(
            InlineKeyboardButton(text=translate_text("Daromad solig'i", lang), callback_data="upload_daromad"),
            InlineKeyboardButton(text=translate_text("Qo‘shilgan qiymat solig‘i", lang), callback_data="upload_qqs")
        )  # Standart tugmalar
        

    logger.info(f"Inline keyboard yaratildi: soliq_turi={soliq_turi}, tugmalar={keyboard.inline_keyboard}")

    # Xabarni yuborish
    try:
        await bot.send_message(user_id, message_text, reply_markup=keyboard, parse_mode="Markdown")
        logger.info(f"Xabar yuborildi: STIR={stir}, tugmalar bilan")
    except Exception as e:
        logger.error(f"Xabar yuborishda xato: {e}")
        await bot.send_message(user_id, translate_text(f"❌ Xabar yuborishda xato: {str(e)}", lang), parse_mode='Markdown')



@dp.callback_query_handler(lambda c: c.data.startswith("upload_"), user_id=ADMIN_IDS)
async def select_month_for_upload(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    soliq_turi = callback_query.data.split("_", 1)[1]
    await state.update_data(soliq_turi=soliq_turi)
    keyboard = InlineKeyboardMarkup(row_width=3)
    oylar = ["yanvar", "fevral", "mart", "aprel", "may", "iyun", "iyul"]
    for oy in oylar:
        keyboard.insert(InlineKeyboardButton(get_month_name(lang, oy), callback_data=f"start_upload_{oy}"))
    await bot.send_message(callback_query.from_user.id, translate_text("Fayllarni qaysi oy uchun yuklamoqchisiz?", lang), reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("start_upload_"), user_id=ADMIN_IDS)
async def start_file_upload(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    oy = callback_query.data.split("_", 2)[2]
    data = await state.get_data()
    stir = data.get('stir')
    soliq_turi = data.get('soliq_turi')
    
    if not stir:
        await bot.send_message(callback_query.from_user.id, translate_text("❌ STIR ma'lumoti yo'q. Iltimos, qayta boshlang.", lang))
        await state.finish()
        return
    await state.update_data(oy=oy)
    
    # Fayllarni tekshirish
    file_path_latin = os.path.normpath(os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_latin', oy)}1.xlsx"))
    file_path_cyrillic = os.path.normpath(os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_cyrillic', oy)}1.xlsx"))
    existing_file = check_file(stir, soliq_turi, oy, "excel1_latin") or check_file(stir, soliq_turi, oy, "excel1_cyrillic")
    
    if existing_file and os.path.exists(existing_file):
        # Agar fayl mavjud bo'lsa, 2-Excel faylini so'rash
        await UploadFiles.excel2.set()
        await bot.send_message(
            callback_query.from_user.id,
            translate_text(f"✅ {get_month_name(lang, oy)} uchun 1-Excel fayl allaqachon mavjud. Endi 2-Excel faylni yuklang (.xlsx):", lang)
        )
    else:
        # Agar fayl mavjud bo'lmasa, 1-Excel faylni so'rash
        await UploadFiles.excel1.set()
        await bot.send_message(
            callback_query.from_user.id,
            translate_text(f"📋 {get_month_name(lang, oy)} uchun 1-Excel faylni yuklang (.xlsx):", lang)
        )

    logger.info(f"start_file_upload: user_id={user_id}, stir={stir}, oy={oy}, soliq_turi={soliq_turi}, existing_file={existing_file}")

    
@dp.callback_query_handler(lambda c: c.data.startswith("overwrite_"), user_id=ADMIN_IDS)
async def overwrite_file(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    oy = callback_query.data.split("_", 1)[1]
    await state.update_data(oy=oy)
    await UploadFiles.excel1.set()
    await bot.send_message(callback_query.from_user.id, translate_text("1-Excel faylni yuklang (.xlsx):", lang))


from database import add_firma, add_firm_owner

# 📌 Telefonni qabul qilish — Bazaga saqlash
@dp.message_handler(state=AddFirma.phone, user_id=ADMIN_IDS)
async def process_phone(message: types.Message, state: FSMContext):
    phone = message.text.strip()

    if not phone.startswith("+998") or len(phone) != 13:
        return await message.answer("❌ Telefon formati noto‘g‘ri!\nMasalan: +998901234567")

    data = await state.get_data()
    stir = data["stir"]
    name = data["name"]
    rahbar = data["rahbar"]
    soliq_turi = data["soliq_turi"]

    add_firma(stir, name, rahbar, soliq_turi)
    add_firm_owner(stir, phone)

    # papka
    os.makedirs(os.path.join(DATA_PATH, stir, "daromad"), exist_ok=True)
    if soliq_turi == "ds-ys":
        os.makedirs(os.path.join(DATA_PATH, stir, "yagona"), exist_ok=True)
    else:
        os.makedirs(os.path.join(DATA_PATH, stir, "qqs"), exist_ok=True)

    await message.answer(
        f"✅ Firma qo‘shildi!\n"
        f"🏢 {name} ({stir})\n"
        f"👤 Rahbar: {rahbar}\n"
        f"📞 Telefon: {phone}\n"
        f"💼 Soliq turi: {soliq_turi}"
    )

    logger.info(f"Yangi firma: {stir}, {name}, {rahbar}, {phone}, {soliq_turi}")
    await state.finish()





def parse_yagona_excel(file_path, lang='uz_latin'):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Лист1']
        firms = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            stir, oy, firma_nomi, rahbar, soliq_turi_yagona, yil_boshidan_aylanma, shu_oy_aylanma = row
            if not stir or not oy or not firma_nomi or not rahbar or not soliq_turi_yagona:
                logger.warning(f"Noto'g'ri qator: {row}")
                continue
            if not re.match(r'^\d{9}$', str(stir)):
                logger.warning(f"Noto'g'ri STIR: {stir}")
                continue
            if not check_firma(str(stir)):
                logger.warning(f"STIR ma'lumotlar bazasida yo'q: {stir}")
                continue
            oy = oy.lower()
            if oy not in ['yanvar', 'fevral', 'mart', 'aprel', 'may', 'iyun', 'iyul']:
                logger.warning(f"Noto'g'ri oy: {oy}")
                continue
            # Standardize soliq_turi_yagona to string with % for consistency
            if isinstance(soliq_turi_yagona, (int, float)):
                soliq_turi_yagona = f"{soliq_turi_yagona}%"
            if lang == 'uz_cyrillic':
                firma_nomi = convert_to_cyrillic(firma_nomi)
                rahbar = convert_to_cyrillic(rahbar)
            else:
                firma_nomi = translate_text(firma_nomi, lang)
                rahbar = translate_text(rahbar, lang)
            key = (str(stir), oy)
            firms[key] = {
                'stir': str(stir),
                'oy': oy,
                'firma_nomi': firma_nomi,
                'rahbar': rahbar,
                'soliq_turi_yagona': soliq_turi_yagona,
                'yil_boshidan_aylanma': int(yil_boshidan_aylanma),
                'shu_oy_aylanma': int(shu_oy_aylanma)
            }
        return firms, None
    except Exception as e:
        logger.error(f"Yagona Excel parsing xatosi: {e}")
        return None, f"Yagona Excel faylni o'qishda xato: {str(e)}"

def parse_qqs_excel(file_path, lang='uz_latin'):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Лист1']  # Excel faylidagi sahifa nomi
        firms = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            stir, oy, firma_nomi, rahbar, soliq_turi_qqs, yil_boshidan_qqs, shu_oy_qqs = row

            if not stir or not oy or not firma_nomi or not rahbar or not soliq_turi_qqs:
                logger.warning(f"Noto'g'ri qator: {row}")
                continue
            if not re.match(r'^\d{9}$', str(stir)):
                logger.warning(f"Noto'g'ri STIR: {stir}")
                continue
            if not check_firma(str(stir)):
                logger.warning(f"STIR ma'lumotlar bazasida yo'q: {stir}")
                continue

            oy = oy.lower()
            if oy not in ['yanvar', 'fevral', 'mart', 'aprel', 'may', 'iyun', 'iyul']:
                logger.warning(f"Noto'g'ri oy: {oy}")
                continue

            # Ma'lumotlarni tilga qarab tarjima qilish
            if lang == 'uz_cyrillic':
                firma_nomi = convert_to_cyrillic(firma_nomi)
                rahbar = convert_to_cyrillic(rahbar)
            else:
                firma_nomi = translate_text(firma_nomi, lang)
                rahbar = translate_text(rahbar, lang)

            key = (str(stir), oy)
            firms[key] = {
                'stir': str(stir),
                'oy': oy,
                'firma_nomi': firma_nomi,
                'rahbar': rahbar,
                'soliq_turi_qqs': soliq_turi_qqs,
                'yil_boshidan_qqs': int(yil_boshidan_qqs),
                'shu_oy_qqs': int(shu_oy_qqs)
            }

        return firms, None
    except Exception as e:
        logger.error(f"QQS Excel parsing xatosi: {e}")
        return None, f"QQS Excel faylni o'qishda xato: {str(e)}"
    


def generate_yagona_excel(stir, oy, firma_nomi, rahbar, soliq_turi_yagona, yil_boshidan_aylanma, shu_oy_aylanma, dest_path_latin, dest_path_cyrillic):
    try:
        # Lotin tilida fayl yaratish
        workbook_latin = openpyxl.Workbook()
        sheet_latin = workbook_latin.active
        sheet_latin.title = "Sheet1"
        headers_latin = [
            translate_text("STIR", 'uz_latin'),
            translate_text("Oy", 'uz_latin'),
            translate_text("Firma nomi", 'uz_latin'),
            translate_text("Raxbar", 'uz_latin'),
            translate_text("Soliq turi yagona", 'uz_latin'),
            translate_text("Yil boshidan aylanma", 'uz_latin'),
            translate_text("Shu oy uchun aylanma", 'uz_latin')
        ]
        sheet_latin.append(headers_latin)
        row = [
            stir,
            get_month_name('uz_latin', oy),
            translate_text(firma_nomi, 'uz_latin'),
            translate_text(rahbar, 'uz_latin'),
            soliq_turi_yagona,
            yil_boshidan_aylanma,
            shu_oy_aylanma
        ]
        sheet_latin.append(row)

        os.makedirs(os.path.dirname(dest_path_latin), exist_ok=True)
        workbook_latin.save(dest_path_latin)
        logger.info(f"Yagona Excel fayli yaratildi (lotin): {dest_path_latin}")

        # Kirill tilida fayl yaratish
        workbook_cyrillic = openpyxl.Workbook()
        sheet_cyrillic = workbook_cyrillic.active
        sheet_cyrillic.title = "Лист1"
        headers_cyrillic = [
            translate_text("STIR", 'uz_cyrillic'),
            translate_text("Oy", 'uz_cyrillic'),
            translate_text("Firma nomi", 'uz_cyrillic'),
            translate_text("Raxbar", 'uz_cyrillic'),
            translate_text("Soliq turi yagona", 'uz_cyrillic'),
            translate_text("Yil boshidan aylanma", 'uz_cyrillic'),
            translate_text("Shu oy uchun aylanma", 'uz_cyrillic')
        ]
        sheet_cyrillic.append(headers_cyrillic)
        row = [
            stir,
            get_month_name('uz_cyrillic', oy),
            translate_text(firma_nomi, 'uz_cyrillic'),
            translate_text(rahbar, 'uz_cyrillic'),
            soliq_turi_yagona,
            yil_boshidan_aylanma,
            shu_oy_aylanma
        ]
        sheet_cyrillic.append(row)

        os.makedirs(os.path.dirname(dest_path_cyrillic), exist_ok=True)
        workbook_cyrillic.save(dest_path_cyrillic)
        logger.info(f"Yagona Excel fayli yaratildi (kirill): {dest_path_cyrillic}")

        return True
    except Exception as e:
        logger.error(f"Yagona Excel faylini yaratishda xato: {e}")
        return False

@dp.message_handler(content_types=['document'], state=UploadFiles.excel1, user_id=ADMIN_IDS)
async def process_excel1(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    if not message.document.file_name.endswith('.xlsx'):
        await message.answer(translate_text("❌ Faqat .xlsx fayllarni yuklang.", lang))
        logger.warning(f"Noto'g'ri fayl formati: {message.document.file_name}")
        return
    data = await state.get_data()
    stir = data['stir']
    soliq_turi = data['soliq_turi']
    oy = data['oy'].lower()

    file_path_latin = os.path.normpath(os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_latin', oy)}1.xlsx"))
    file_path_cyrillic = os.path.normpath(os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_cyrillic', oy)}1.xlsx"))
    os.makedirs(os.path.dirname(file_path_latin), exist_ok=True)

    temp_path = os.path.normpath(os.path.join(DATA_PATH, "temp", f"excel1_{user_id}_{int(datetime.now().timestamp())}.xlsx"))
    os.makedirs(os.path.dirname(temp_path), exist_ok=True)

    try:
        await message.document.download(destination_file=temp_path)
        if soliq_turi == 'yagona':
            firms, error = parse_yagona_excel(temp_path, lang)
            if error or not firms:
                await message.answer(translate_text(f"❌ Faylni o'qishda xato: {error}", lang))
                logger.error(f"Yagona faylni o'qishda xato: {error}, temp_path={temp_path}")
                return
            key = (stir, oy)
            if key in firms:
                firm = firms[key]
                generate_yagona_excel(
                    stir, oy, firm['firma_nomi'], firm['rahbar'], firm['soliq_turi_yagona'],
                    firm['yil_boshidan_aylanma'], firm['shu_oy_aylanma'], file_path_latin, file_path_cyrillic
                )
        elif soliq_turi == 'qqs':
            firms, error = parse_qqs_excel(temp_path, lang)
            if error or not firms:
                await message.answer(translate_text(f"❌ Faylni o'qishda xato: {error}", lang))
                logger.error(f"QQS faylni o'qishda xato: {error}, temp_path={temp_path}")
                return
            import shutil
            shutil.copy(temp_path, file_path_latin)
            shutil.copy(temp_path, file_path_cyrillic)
        else:
            firms, error = parse_excel_file(temp_path, lang)
            if error or not firms:
                await message.answer(translate_text(f"❌ Faylni o'qishda xato: {error}", lang))
                logger.error(f"Daromad faylni o'qishda xato: {error}, temp_path={temp_path}")
                return
            import shutil
            shutil.copy(temp_path, file_path_latin)
            shutil.copy(temp_path, file_path_cyrillic)

        save_file(stir, soliq_turi, oy, "excel1_latin", file_path_latin)
        save_file(stir, soliq_turi, oy, "excel1_cyrillic", file_path_cyrillic)
        logger.info(f"Fayl yuklandi: {file_path_latin}, {file_path_cyrillic}")

        if os.path.exists(temp_path):
            os.remove(temp_path)
            logger.info(f"Vaqtinchalik fayl o'chirildi: {temp_path}")

        await state.update_data(excel_file_path=temp_path)  # Vaqtinchalik fayl yo‘lini saqlash
        await UploadFiles.excel2.set()
        sent_message = await message.answer(translate_text("✅ 1-Excel fayl yuklangan, endi 2-Excel faylni yuklang (.xlsx). Bekor qilish uchun /cancel bosing.", lang))
        await state.update_data(last_message_id=sent_message.message_id)
        logger.info(f"excel1 holatidan excel2 holatiga o'tildi: user_id={user_id}, stir={stir}, oy={oy}")
    except Exception as e:
        logger.error(f"Excel1 faylini yuklashda xato: {e}, user_id={user_id}, temp_path={temp_path}")
        await message.answer(translate_text(f"❌ Faylni yuklashda xato yuz berdi: {str(e)}", lang))

@dp.message_handler(content_types=['document'], state=UploadFiles.excel2, user_id=ADMIN_IDS)
async def process_excel2(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    if not message.document.file_name.endswith('.xlsx'):
        await message.answer(translate_text("❌ Faqat .xlsx fayllarni yuklang.", lang))
        logger.warning(f"Noto'g'ri fayl formati: {message.document.file_name}, user_id={user_id}")
        return
    data = await state.get_data()
    stir = data.get('stir')
    soliq_turi = data.get('soliq_turi')
    oy = data.get('oy').lower()
    
    if not all([stir, soliq_turi, oy]):
        await message.answer(translate_text("❌ STIR, soliq turi yoki oy ma'lumotlari yo'q. Qayta boshlang.", lang))
        logger.error(f"Not enough data: stir={stir}, soliq_turi={soliq_turi}, oy={oy}")
        await state.finish()
        return
    
    file_path_latin = os.path.normpath(os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_latin', oy)}2.xlsx"))
    file_path_cyrillic = os.path.normpath(os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_cyrillic', oy)}2.xlsx"))
    os.makedirs(os.path.dirname(file_path_latin), exist_ok=True)
    
    temp_path = os.path.normpath(os.path.join(DATA_PATH, "temp", f"excel2_{user_id}_{int(datetime.now().timestamp())}.xlsx"))
    os.makedirs(os.path.dirname(temp_path), exist_ok=True)
    
    try:
        await message.document.download(destination_file=temp_path)
        import shutil
        shutil.copy(temp_path, file_path_latin)
        shutil.copy(temp_path, file_path_cyrillic)
        logger.info(f"Fayllar nusxalandi: temp={temp_path}, latin={file_path_latin}, cyrillic={file_path_cyrillic}")
        
        save_file(stir, soliq_turi, oy, "excel2_latin", file_path_latin)
        save_file(stir, soliq_turi, oy, "excel2_cyrillic", file_path_cyrillic)
        logger.info(f"Fayl yuklandi: {file_path_latin}, {file_path_cyrillic}")
        
        if os.path.exists(temp_path):
            os.remove(temp_path)
            logger.info(f"Vaqtinchalik fayl o'chirildi: {temp_path}")
        
        await state.update_data(excel_file_path=temp_path)  # Vaqtinchalik fayl yo‘lini saqlash
        await UploadFiles.next()
        sent_message = await message.answer(translate_text("html_xlsx faylni yuklang yoki /cancel bosib amaliyotni bekor qilin", lang))
        await state.update_data(last_message_id=sent_message.message_id)
        logger.info(f"excel2 holatidan html holatiga o'tildi: user_id={user_id}, stir={stir}, oy={oy}")
    except Exception as e:
        logger.error(f"Excel2 faylini yuklashda xato: {e}, user_id={user_id}, temp_path={temp_path}")
        await message.answer(translate_text(f"❌ Faylni yuklashda xato yuz berdi: {str(e)}", lang))

@dp.message_handler(content_types=['document'], state=UploadFiles.html, user_id=ADMIN_IDS)
async def process_html(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    logger.info(f"process_html boshlandi: user_id={user_id}, file_name={message.document.file_name}")

    if not message.document.file_name.endswith('.html'):
        await message.answer(translate_text("❌ Faqat .html fayllarni yuklang.", lang), parse_mode='Markdown')
        logger.warning(f"Noto'g'ri fayl formati: {message.document.file_name}, user_id={user_id}")
        return

    data = await state.get_data()
    stir = data.get('stir')
    soliq_turi = data.get('soliq_turi')
    oy = data.get('oy')

    if not all([stir, soliq_turi, oy, user_id]):
        await message.answer(translate_text("❌ STIR, soliq turi, oy yoki user_id ma'lumotlari yo'q. Qayta boshlang.", lang), parse_mode='Markdown')
        logger.error(f"Not enough data: stir={stir}, soliq_turi={soliq_turi}, oy={oy}, user_id={user_id}")
        await state.finish()
        return

    oy = oy.lower()
    file_path_latin = os.path.normpath(os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_latin', oy)}3.html"))
    file_path_cyrillic = os.path.normpath(os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_cyrillic', oy)}3.html"))
    os.makedirs(os.path.dirname(file_path_latin), exist_ok=True)
    
    temp_path = os.path.normpath(os.path.join(DATA_PATH, "temp", f"html_{user_id}_{int(datetime.now().timestamp())}.html"))
    os.makedirs(os.path.dirname(temp_path), exist_ok=True)

    try:
        await message.document.download(destination_file=temp_path)
        import shutil
        shutil.copy(temp_path, file_path_latin)
        shutil.copy(temp_path, file_path_cyrillic)
        logger.info(f"Fayllar nusxalandi: temp={temp_path}, latin={file_path_latin}, cyrillic={file_path_cyrillic}")

        save_file(stir, soliq_turi, oy, "html", file_path_latin)
        logger.info(f"Fayl yuklandi: {file_path_latin}, user_id={user_id}")

        if os.path.exists(temp_path):
            os.remove(temp_path)
            logger.info(f"Vaqtinchalik fayl o'chirildi: {temp_path}")

        await state.finish()
        sent_message = await message.answer(translate_text(f"✅ {get_month_name(lang, oy)} uchun fayllar muvaffaqiyatli yuklandi!", lang), parse_mode='Markdown')
        await state.update_data(last_message_id=sent_message.message_id)
        logger.info(f"HTML fayl yuklandi va holat yakunlandi: user_id={user_id}, stir={stir}, oy={oy}")
        await back_to_admin_panel(state=state)
    except Exception as e:
        logger.error(f"HTML faylini yuklashda xato: {e}, user_id={user_id}, temp_path={temp_path}")
        await message.answer(translate_text(f"❌ Faylni yuklashda xato yuz berdi: {str(e)}", lang), parse_mode='Markdown')


@dp.callback_query_handler(lambda c: c.data == "delete_report", user_id=ADMIN_IDS)
async def start_delete_report(callback_query: types.CallbackQuery):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    firms = get_all_firms()
    if not firms:
        await bot.send_message(callback_query.from_user.id, translate_text("❌ Hozircha firmalar mavjud emas.", lang))
        return
    keyboard, page, total_pages = create_paginated_keyboard(firms, "delete_firm", page=1, lang=lang)
    await bot.send_message(callback_query.from_user.id, translate_text(f"Hisobotni o'chirish uchun firma tanlang (Sahifa {page}/{total_pages}):", lang), reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("delete_firm_page_"), user_id=ADMIN_IDS)
async def delete_firma_paginate(callback_query: types.CallbackQuery):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    page = int(callback_query.data.split("_")[-1])
    firms = get_all_firms()
    keyboard, page, total_pages = create_paginated_keyboard(firms, "delete_firm", page=page, lang=lang)
    await bot.edit_message_text(
        translate_text(f"Hisobotni o'chirish uchun firma tanlang (Sahifa {page}/{total_pages}):", lang),
        callback_query.from_user.id,
        callback_query.message.message_id,
        reply_markup=keyboard
    )

@dp.callback_query_handler(lambda c: c.data == "delete_firm_search", user_id=ADMIN_IDS, state="*")
async def start_delete_firma_search(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    await state.finish()
    await ManualInput.search.set()
    await state.update_data(search_context="delete_report")
    await bot.send_message(callback_query.from_user.id, translate_text("Firma STIR yoki nomini kiriting (qisman moslik uchun):", lang))

@dp.callback_query_handler(lambda c: c.data.startswith("delete_firm_"), user_id=ADMIN_IDS)
async def select_month_to_delete(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    stir = callback_query.data.split("_", 2)[2]
    await state.update_data(stir=stir)
    keyboard = InlineKeyboardMarkup(row_width=3)
    oylar = ["yanvar", "fevral", "mart", "aprel", "may", "iyun", "iyul"]
    for oy in oylar:
        keyboard.insert(InlineKeyboardButton(get_month_name(lang, oy), callback_data=f"delete_oy_{stir}_{oy}"))
    await bot.send_message(callback_query.from_user.id, translate_text("Qaysi oyning hisobotini o'chirishni xohlaysiz?", lang), reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("delete_oy_"), user_id=ADMIN_IDS)
async def confirm_delete_report(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    _, _, stir, oy = callback_query.data.split("_")
    await state.update_data(oy=oy)
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Ha, o'chirish", lang), callback_data=f"confirm_delete_{stir}_{oy}"),
        InlineKeyboardButton(translate_text("Yo'q, bekor qilish", lang), callback_data="cancel_delete")
    )
    await bot.send_message(callback_query.from_user.id, translate_text(f"{get_month_name(lang, oy)} oyi uchun hisobotni o'chirishni xohlaysizmi?", lang), reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("confirm_delete_"), user_id=ADMIN_IDS)
async def delete_report(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    _, _, stir, oy = callback_query.data.split("_")
    
    # Firma soliq turini olish
    conn = sqlite3.connect(os.path.join(DATA_PATH, "bot.db"))
    c = conn.cursor()
    c.execute("SELECT soliq_turi FROM firms WHERE stir = ?", (stir,))
    result = c.fetchone()
    soliq_turi = result[0].lower() if result and result[0] else 'daromad'
    
    # Hisobot va fayllarni o'chirish
    c.execute("DELETE FROM reports WHERE stir = ? AND oy = ?", (stir, oy))
    c.execute("DELETE FROM files WHERE stir = ? AND oy = ?", (stir, oy))
    conn.commit()
    conn.close()
    
    # Faqat firma soliq turiga mos fayllarni o'chirish
    for file_type in ["excel1_latin", "excel1_cyrillic", "excel2_latin", "excel2_cyrillic", "html"]:
        file_path_latin = os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_latin', oy)}{'1' if 'excel1' in file_type else '2' if 'excel2' in file_type else '3'}{'' if file_type == 'html' else '.xlsx'}")
        file_path_cyrillic = os.path.join(DATA_PATH, stir, soliq_turi, f"{get_month_name('uz_cyrillic', oy)}{'1' if 'excel1' in file_type else '2' if 'excel2' in file_type else '3'}{'' if file_type == 'html' else '.xlsx'}")
        for file_path in [file_path_latin, file_path_cyrillic]:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Fayl o'chirildi: {file_path}")
    
    await state.finish()
    await bot.send_message(callback_query.from_user.id, translate_text(f"✅ {get_month_name(lang, oy)} oyi hisoboti o'chirildi.", lang))
    logger.info(f"Hisobot o'chirildi: STIR={stir}, Oy={oy}")


@dp.callback_query_handler(lambda c: c.data == "cancel_delete", user_id=ADMIN_IDS)
async def cancel_delete(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    await state.finish()
    await bot.send_message(callback_query.from_user.id, translate_text("❌ O'chirish bekor qilindi.", lang))

@dp.callback_query_handler(lambda c: c.data == "manual_input", user_id=ADMIN_IDS)
async def start_manual_input(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Daromad solig'i", lang), callback_data="manual_daromad"),
        InlineKeyboardButton(translate_text("Yagona soliq", lang), callback_data="manual_yagona"),
        InlineKeyboardButton(translate_text("Qo‘shilgan qiymat solig‘i", lang), callback_data="manual_qqs")
    )
    await ManualInput.select_soliq_turi.set()
    await bot.send_message(
        callback_query.from_user.id,
        translate_text("Hisobot kiritish uchun soliq turini tanlang:", lang),
        reply_markup=keyboard
    )

@dp.callback_query_handler(lambda c: c.data in ["manual_daromad", "manual_yagona", "manual_qqs"], user_id=ADMIN_IDS, state=ManualInput.select_soliq_turi)
async def process_soliq_turi_selection(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    soliq_turi = callback_query.data.split("_")[1]  # "daromad", "yagona", or "qqs"
    await state.update_data(soliq_turi=soliq_turi)

    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Excel fayl yuklash", lang), callback_data="upload_excel"),
        InlineKeyboardButton(translate_text("Qo'lda kiritish", lang), callback_data="manual_no_excel")
    )
    await ManualInput.excel_upload.set()
    await bot.send_message(
        callback_query.from_user.id,
        translate_text(f"{translate_text(soliq_turi.capitalize() + ' solig‘i', lang)} hisobotini kiritish usulini tanlang:", lang),
        reply_markup=keyboard
    )



@dp.callback_query_handler(lambda c: c.data == "upload_excel", user_id=ADMIN_IDS, state=ManualInput.excel_upload)
async def request_excel_file(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    last_message_id = data.get('last_message_id')

    # Avvalgi xabarni o‘chirish
    if last_message_id:
        try:
            await bot.delete_message(chat_id=user_id, message_id=last_message_id)
        except Exception as e:
            logger.warning(f"Xabar o'chirishda xato: {e}, message_id={last_message_id}")

    sent_message = await bot.send_message(
        callback_query.from_user.id,
        translate_text("xlsx faylni yuklang yoki /cancel bosib amaliyotni bekor qilin", lang)
    )
    await state.update_data(last_message_id=sent_message.message_id)
@dp.message_handler(content_types=['document'], state=ManualInput.excel_upload, user_id=ADMIN_IDS)
async def process_excel_upload(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    if not message.document.file_name.endswith('.xlsx'):
        await message.answer(translate_text("❌ Faqat .xlsx fayllarni yuklang.", lang))
        return
    data = await state.get_data()
    soliq_turi = data.get('soliq_turi')
    file_path = os.path.join(DATA_PATH, "temp", f"manual_{user_id}_{int(datetime.now().timestamp())}.xlsx")
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    await message.document.download(destination_file=file_path)

    if soliq_turi == 'daromad':
        firms, error = parse_excel_file(file_path, lang)
    elif soliq_turi == 'yagona':
        firms, error = parse_yagona_excel(file_path, lang)
    elif soliq_turi == 'qqs':
        firms, error = parse_qqs_excel(file_path, lang)
    else:
        await message.answer(translate_text("❌ Noto'g'ri soliq turi.", lang))
        return

    if not firms:
        await message.answer(translate_text(f"❌ Faylni o'qishda xatolik yuz berdi: {error}", lang))
        if os.path.exists(file_path):
            os.remove(file_path)
        return

    await state.update_data(excel_file_path=file_path, firms=firms)
    firms_list = [(k[0], k[1], v['firma_nomi']) for k, v in firms.items()]
    keyboard, page, total_pages = create_paginated_keyboard(firms_list, "manual_firm", page=1, lang=lang)
    await ManualInput.stir.set()
    await message.answer(
        translate_text(f"Excel faylidan quyidagi firmalar topildi. Hisobot kiritish uchun birini tanlang (Sahifa {page}/{total_pages}):", lang),
        reply_markup=keyboard
    )


@dp.callback_query_handler(lambda c: c.data.startswith("manual_firm_page_"), user_id=ADMIN_IDS, state=ManualInput.stir)
async def manual_firm_paginate(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    page = int(callback_query.data.split("_")[-1])
    data = await state.get_data()
    firms = data.get('firms', {})
    firms_list = [(k[0], k[1], v['firma_nomi']) for k, v in firms.items()]
    keyboard, page, total_pages = create_paginated_keyboard(firms_list, "manual_firm", page=page, lang=lang)
    await bot.edit_message_text(
        translate_text(f"Excel faylidan quyidagi firmalar topildi. Hisobot kiritish uchun birini tanlang (Sahifa {page}/{total_pages}):", lang),
        callback_query.from_user.id,
        callback_query.message.message_id,
        reply_markup=keyboard
    )

@dp.callback_query_handler(lambda c: c.data == "manual_firm_search", user_id=ADMIN_IDS, state=ManualInput.stir)
async def start_manual_firm_search(callback_query: types.CallbackQuery, state=FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    await ManualInput.search.set()
    await state.update_data(search_context="manual_excel")
    await bot.send_message(callback_query.from_user.id, translate_text("Excel faylidagi firma STIR yoki nomini kiriting (qisman moslik uchun):", lang))

@dp.callback_query_handler(lambda c: c.data == "manual_no_excel", user_id=ADMIN_IDS, state=ManualInput.excel_upload)
async def skip_excel_upload(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    soliq_turi = data.get('soliq_turi')

    firms = get_all_firms()
    if not firms:
        await bot.send_message(callback_query.from_user.id, translate_text("❌ Hozircha firmalar mavjud emas.", lang))
        await state.finish()
        return

    keyboard, page, total_pages = create_paginated_keyboard(firms, "manual_firm", page=1, lang=lang)
    await ManualInput.stir.set()
    await bot.send_message(
        callback_query.from_user.id,
        translate_text(f"Hisobot kiritish uchun firmani tanlang (Sahifa {page}/{total_pages}):", lang),
        reply_markup=keyboard
    )
@dp.callback_query_handler(lambda c: c.data.startswith("manual_firm_"), user_id=ADMIN_IDS, state=ManualInput.stir)
async def select_firma_or_month(callback_query: types.CallbackQuery, state=FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    parts = callback_query.data.split("_")
    stir = parts[2]
    oy = parts[3] if len(parts) > 3 else None
    soliq_turi = data.get('soliq_turi')

    if not re.match(r'^\d{9}$', stir):
        await bot.send_message(callback_query.from_user.id, translate_text("❌ Noto'g'ri STIR formati. /admin orqali qayta boshlang.", lang))
        return

    await state.update_data(stir=stir)

    if oy:
        if oy not in ['yanvar', 'fevral', 'mart', 'aprel', 'may', 'iyun', 'iyul']:
            await bot.send_message(callback_query.from_user.id, translate_text("❌ Noto'g'ri oy formati. /admin orqali qayta boshlang.", lang))
            return
        await state.update_data(oy=oy)
        if soliq_turi == 'daromad':
            await process_excel_data(callback_query, state)
        elif soliq_turi == 'yagona':
            firms = data.get('firms', {})
            key = (stir, oy)
            if key in firms:
                firm = firms[key]
                try:
                    # Handle soliq_turi_yagona as string or float
                    soliq_turi_yagona = firm['soliq_turi_yagona']
                    if isinstance(soliq_turi_yagona, str):
                        soliq_rate = float(soliq_turi_yagona.strip('%')) / 100
                    elif isinstance(soliq_turi_yagona, (int, float)):
                        soliq_rate = soliq_turi_yagona / 100
                    else:
                        raise ValueError(f"Invalid soliq_turi_yagona format: {soliq_turi_yagona}")
                    
                    yagona_soliq = int(firm['shu_oy_aylanma'] * soliq_rate)
                    result = get_text(
                        lang,
                        'yagona_report',
                        firma_nomi=firm['firma_nomi'],
                        rahbar=firm['rahbar'],
                        oy=get_month_name(lang, oy),
                        yil_boshidan_aylanma=f"{firm['yil_boshidan_aylanma']:,}",
                        shu_oy_aylanma=f"{firm['shu_oy_aylanma']:,}",
                        soliq_turi_yagona=soliq_turi_yagona,
                        yagona_soliq=f"{yagona_soliq:,}"
                    )
                    # Save firma_name to state
                    await state.update_data(
                        firma_name=firm['firma_nomi'],  # Added to fix KeyError
                        soliq_turi_yagona=soliq_turi_yagona,
                        yil_boshidan_aylanma=firm['yil_boshidan_aylanma'],
                        shu_oy_aylanma=firm['shu_oy_aylanma'],
                        yagona_soliq=yagona_soliq,
                        rahbar=firm['rahbar']
                    )
                    keyboard = InlineKeyboardMarkup(row_width=2)
                    keyboard.add(
                        InlineKeyboardButton(translate_text("Tasdiqlash", lang), callback_data="confirm_report"),
                        InlineKeyboardButton(translate_text("Tahrirlash", lang), callback_data="edit_report"),
                        InlineKeyboardButton(translate_text("Bekor qilish", lang), callback_data="cancel_report")
                    )
                    await ManualInput.confirm.set()
                    await bot.send_message(callback_query.from_user.id, result + "\n" + translate_text("Tasdiqlaysizmi?", lang), reply_markup=keyboard)
                except Exception as e:
                    logger.error(f"Yagona soliq hisoblashda xato: STIR={stir}, Oy={oy}, Error={str(e)}")
                    await bot.send_message(callback_query.from_user.id, translate_text(f"❌ Yagona soliq hisoblashda xato: {str(e)}", lang))
                    return
            else:
                firma_name = get_firma_name(stir)
                await state.update_data(firma_name=firma_name)
                await ManualInput.yagona_data.set()
                await bot.send_message(
                    callback_query.from_user.id,
                    translate_text(
                        f"Excel faylida {stir} uchun {get_month_name(lang, oy)} ma'lumotlari topilmadi.\n"
                        f"Yagona soliq ma'lumotlarini kiriting (soliq stavkasi %, yil boshidan aylanma, shu oy aylanma, masalan: 4%, 10000000, 5000000):",
                        lang
                    )
                )
        elif soliq_turi == 'qqs':
            firms = data.get('firms', {})
            key = (stir, oy)
            if key in firms:
                firm = firms[key]
                try:
                    # Handle soliq_turi_qqs as string or float
                    soliq_turi_qqs = firm['soliq_turi_qqs']
                    if isinstance(soliq_turi_qqs, str):
                        soliq_rate = float(soliq_turi_qqs.strip('%')) / 100
                    elif isinstance(soliq_turi_qqs, (int, float)):
                        soliq_rate = soliq_turi_qqs / 100
                    else:
                        raise ValueError(f"Invalid soliq_turi_qqs format: {soliq_turi_qqs}")
                    
                    qqs_soliq = int(firm['shu_oy_qqs'] * soliq_rate)
                    result = get_text(
                        lang,
                        'qqs_report',
                        firma_nomi=firm['firma_nomi'],
                        rahbar=firm['rahbar'],
                        oy=get_month_name(lang, oy),
                        yil_boshidan_qqs=f"{firm['yil_boshidan_qqs']:,}",
                        shu_oy_qqs=f"{firm['shu_oy_qqs']:,}",
                        soliq_turi_qqs=soliq_turi_qqs,
                        qqs_soliq=f"{qqs_soliq:,}"
                    )
                    # Save firma_name to state
                    await state.update_data(
                        firma_name=firm['firma_nomi'],  # Added to fix KeyError
                        soliq_turi_qqs=soliq_turi_qqs,
                        yil_boshidan_qqs=firm['yil_boshidan_qqs'],
                        shu_oy_qqs=firm['shu_oy_qqs'],
                        qqs_soliq=qqs_soliq,
                        rahbar=firm['rahbar']
                    )
                    keyboard = InlineKeyboardMarkup(row_width=2)
                    keyboard.add(
                        InlineKeyboardButton(translate_text("Tasdiqlash", lang), callback_data="confirm_report"),
                        InlineKeyboardButton(translate_text("Tahrirlash", lang), callback_data="edit_report"),
                        InlineKeyboardButton(translate_text("Bekor qilish", lang), callback_data="cancel_report")
                    )
                    await ManualInput.confirm.set()
                    await bot.send_message(callback_query.from_user.id, result + "\n" + translate_text("Tasdiqlaysizmi?", lang), reply_markup=keyboard)
                except Exception as e:
                    logger.error(f"QQS soliq hisoblashda xato: STIR={stir}, Oy={oy}, Error={str(e)}")
                    await bot.send_message(callback_query.from_user.id, translate_text(f"❌ QQS soliq hisoblashda xato: {str(e)}", lang))
                    return
            else:
                firma_name = get_firma_name(stir)
                await state.update_data(firma_name=firma_name)
                await ManualInput.qqs_data.set()
                await bot.send_message(
                    callback_query.from_user.id,
                    translate_text(
                        f"Excel faylida {stir} uchun {get_month_name(lang, oy)} ma'lumotlari topilmadi.\n"
                        f"QQS ma'lumotlarini kiriting (soliq stavkasi %, yil boshidan QQS, shu oy QQS, masalan: 15%, 20000000, 10000000):",
                        lang
                    )
                )
    else:
        keyboard = InlineKeyboardMarkup(row_width=3)
        oylar = ["yanvar", "fevral", "mart", "aprel", "may", "iyun", "iyul"]
        for oy in oylar:
            keyboard.insert(InlineKeyboardButton(get_month_name(lang, oy), callback_data=f"manual_oy_{stir}_{oy}"))
        await bot.send_message(
            callback_query.from_user.id,
            translate_text("Qaysi oy uchun hisobot kiritmoqchisiz?", lang),
            reply_markup=keyboard
        )

@dp.callback_query_handler(lambda c: c.data.startswith("manual_oy_"), user_id=ADMIN_IDS, state=ManualInput.stir)
async def select_month_manual(callback_query: types.CallbackQuery, state=FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    try:
        parts = callback_query.data.split("_")
        if len(parts) != 4:
            raise ValueError(f"Noto'g'ri callback format: {callback_query.data}")
        stir = parts[2]
        oy = parts[3]
        if not re.match(r'^\d{9}$', stir):
            await bot.send_message(callback_query.from_user.id, translate_text("❌ Noto'g'ri STIR formati. /admin orqali qayta boshlang.", lang))
            return
        if oy not in ['yanvar', 'fevral', 'mart', 'aprel', 'may', 'iyun', 'iyul']:
            await bot.send_message(callback_query.from_user.id, translate_text("❌ Noto'g'ri oy formati. /admin orqali qayta boshlang.", lang))
            return
        await state.update_data(stir=stir, oy=oy)
        await process_excel_data(callback_query, state)
    except Exception as e:
        logger.error(f"select_month_manual xatosi: {e}, callback_data={callback_query.data}")
        await bot.send_message(callback_query.from_user.id, translate_text("❌ Xatolik yuz berdi. /admin orqali qayta boshlang.", lang))

async def process_excel_data(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    stir = data.get('stir')
    oy = data.get('oy')
    firms = data.get('firms', {})
    key = (stir, oy)

    if key in firms:
        firm = firms[key]
        firma_nomi = firm['firma_nomi']
        xodimlar = firm['xodimlar']
        xodimlar_soni = len(xodimlar)
        hisobot_davri_oylik = sum(x['shu_oy'] for x in xodimlar if x['shu_oy'] > 0)
        jami_oylik = sum(x['yil_boshidan'] for x in xodimlar)
        soliq = int(hisobot_davri_oylik * 0.12)

        # Xodimlar ma'lumotlarini to'g'ri formatda shakllantirish
        xodimlar_data = [
            f"{i+1} ({x['lavozim']}) – {x['ism']}, "
            f"{translate_text('bu_oy_uchun_hisobotda', lang)}: {x['shu_oy']:,} {translate_text('so‘m', lang)} "
            f"({translate_text('yil_boshidan_hisobotda', lang)}: {x['yil_boshidan']:,} {translate_text('so‘m', lang)})"
            for i, x in enumerate(xodimlar)
        ]

        result = get_text(lang, 'daromad_report',
                          firma_name=firma_nomi,
                          oy=get_month_name(lang, oy),
                          xodimlar_soni=xodimlar_soni,
                          xodimlar_data='\n'.join(xodimlar_data),
                          jami_oylik=jami_oylik,
                          hisobot_davri_oylik=hisobot_davri_oylik,
                          soliq=soliq)
        await state.update_data(
            firma_name=firma_nomi,
            xodimlar_soni=xodimlar_soni,
            xodimlar_data=xodimlar_data,
            hisobot_davri_oylik=hisobot_davri_oylik,
            jami_oylik=jami_oylik,
            soliq=soliq,
            xodimlar=xodimlar
        )
        keyboard = InlineKeyboardMarkup(row_width=2)
        keyboard.add(
            InlineKeyboardButton(translate_text("Tasdiqlash", lang), callback_data="confirm_report"),
            InlineKeyboardButton(translate_text("Tahrirlash", lang), callback_data="edit_report"),
            InlineKeyboardButton(translate_text("Bekor qilish", lang), callback_data="cancel_report")
        )
        await ManualInput.confirm.set()
        await bot.send_message(callback_query.from_user.id, result + "\n" + translate_text("Tasdiqlaysizmi?", lang), reply_markup=keyboard)
    else:
        firma_name = get_firma_name(stir)
        await state.update_data(firma_name=firma_name)
        await ManualInput.firma_name.set()
        await bot.send_message(callback_query.from_user.id, translate_text(f"Excel faylida {stir} uchun {get_month_name(lang, oy)} ma'lumotlari topilmadi.\nFirma nomi (hozirgi: {firma_name}, o'zgartirish uchun yangi nom kiriting yoki bo'sh qoldiring):", lang))


@dp.message_handler(state=ManualInput.yagona_data, user_id=ADMIN_IDS)
async def process_yagona_data(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    stir = data.get('stir')
    oy = data.get('oy')
    firma_name = data.get('firma_name')

    # Yagona soliq ma'lumotlari uchun format: soliq_turi_yagona, yil_boshidan_aylanma, shu_oy_aylanma
    pattern = r'^\s*([\d.]+%)\s*,\s*([\d\s]+)\s*,\s*([\d\s]+)\s*$'
    match = re.match(pattern, message.text.strip())

    if not match:
        await message.answer(
            translate_text(
                f"❌ Noto'g'ri format. Namuna: 4%, 10000000, 5000000\n"
                f"Kiritilgan matn: {message.text.strip()}",
                lang
            )
        )
        logger.error(f"Noto'g'ri yagona format kiritildi: user_id={user_id}, matn={message.text.strip()}")
        return

    soliq_turi_yagona, yil_boshidan_aylanma, shu_oy_aylanma = match.groups()
    try:
        yil_boshidan_aylanma = int(yil_boshidan_aylanma.replace(" ", ""))
        shu_oy_aylanma = int(shu_oy_aylanma.replace(" ", ""))
    except ValueError:
        await message.answer(translate_text("❌ Aylanma summalari raqam bo'lishi kerak.", lang))
        return

    yagona_soliq = int(shu_oy_aylanma * (float(soliq_turi_yagona.strip('%')) / 100))
    rahbar = get_firma_info(stir)[1] or "Noma'lum"

    result = get_text(
        lang,
        'yagona_report',
        firma_nomi=firma_name,
        rahbar=rahbar,
        oy=get_month_name(lang, oy),
        yil_boshidan_aylanma=f"{yil_boshidan_aylanma:,}",
        shu_oy_aylanma=f"{shu_oy_aylanma:,}",
        soliq_turi_yagona=soliq_turi_yagona,
        yagona_soliq=f"{yagona_soliq:,}"
    )
    await state.update_data(
        soliq_turi_yagona=soliq_turi_yagona,
        yil_boshidan_aylanma=yil_boshidan_aylanma,
        shu_oy_aylanma=shu_oy_aylanma,
        yagona_soliq=yagona_soliq,
        rahbar=rahbar
    )
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Tasdiqlash", lang), callback_data="confirm_report"),
        InlineKeyboardButton(translate_text("Tahrirlash", lang), callback_data="edit_report"),
        InlineKeyboardButton(translate_text("Bekor qilish", lang), callback_data="cancel_report")
    )
    await ManualInput.confirm.set()
    await message.answer(result + "\n" + translate_text("Tasdiqlaysizmi?", lang), reply_markup=keyboard)

@dp.message_handler(state=ManualInput.qqs_data, user_id=ADMIN_IDS)
async def process_qqs_data(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    stir = data.get('stir')
    oy = data.get('oy')
    firma_name = data.get('firma_name')

    # QQS ma'lumotlari uchun format: soliq_turi_qqs, yil_boshidan_qqs, shu_oy_qqs
    pattern = r'^\s*([\d.]+%)\s*,\s*([\d\s]+)\s*,\s*([\d\s]+)\s*$'
    match = re.match(pattern, message.text.strip())

    if not match:
        await message.answer(
            translate_text(
                f"❌ Noto'g'ri format. Namuna: 15%, 20000000, 10000000\n"
                f"Kiritilgan matn: {message.text.strip()}",
                lang
            )
        )
        logger.error(f"Noto'g'ri QQS format kiritildi: user_id={user_id}, matn={message.text.strip()}")
        return

    soliq_turi_qqs, yil_boshidan_qqs, shu_oy_qqs = match.groups()
    try:
        yil_boshidan_qqs = int(yil_boshidan_qqs.replace(" ", ""))
        shu_oy_qqs = int(shu_oy_qqs.replace(" ", ""))
    except ValueError:
        await message.answer(translate_text("❌ QQS summalari raqam bo'lishi kerak.", lang))
        return

    qqs_soliq = int(shu_oy_qqs * (float(soliq_turi_qqs.strip('%')) / 100))
    rahbar = get_firma_info(stir)[1] or "Noma'lum"

    result = get_text(
        lang,
        'qqs_report',
        firma_nomi=firma_name,
        rahbar=rahbar,
        oy=get_month_name(lang, oy),
        yil_boshidan_qqs=f"{yil_boshidan_qqs:,}",
        shu_oy_qqs=f"{shu_oy_qqs:,}",
        soliq_turi_qqs=soliq_turi_qqs,
        qqs_soliq=f"{qqs_soliq:,}"
    )
    await state.update_data(
        soliq_turi_qqs=soliq_turi_qqs,
        yil_boshidan_qqs=yil_boshidan_qqs,
        shu_oy_qqs=shu_oy_qqs,
        qqs_soliq=qqs_soliq,
        rahbar=rahbar
    )
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Tasdiqlash", lang), callback_data="confirm_report"),
        InlineKeyboardButton(translate_text("Tahrirlash", lang), callback_data="edit_report"),
        InlineKeyboardButton(translate_text("Bekor qilish", lang), callback_data="cancel_report")
    )
    await ManualInput.confirm.set()
    await message.answer(result + "\n" + translate_text("Tasdiqlaysizmi?", lang), reply_markup=keyboard)




@dp.message_handler(state=ManualInput.firma_name, user_id=ADMIN_IDS)
async def process_firma_name(message: types.Message, state=FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    firma_name = message.text.strip()
    data = await state.get_data()
    stir = data['stir']
    if not firma_name:
        firma_name = data.get('firma_name', get_firma_name(stir))
    if len(firma_name) < 3:
        await message.answer(translate_text("❌ Firma nomi kamida 3 ta belgidan iborat bo'lishi kerak.", lang))
        return
    if lang == 'uz_cyrillic':
        firma_name = convert_to_cyrillic(firma_name)
    await state.update_data(firma_name=firma_name)
    await ManualInput.xodimlar_soni.set()
    await message.answer(translate_text(f"Xodimlar sonini kiriting (raqam bilan, masalan: 2):", lang))

@dp.message_handler(state=ManualInput.xodimlar_soni, user_id=ADMIN_IDS)
async def process_xodimlar_soni(message: types.Message, state=FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    try:
        xodimlar_soni = int(message.text.strip())
        if xodimlar_soni <= 0:
            await message.answer(translate_text("❌ Xodimlar soni 0 dan katta bo'lishi kerak.", lang))
            return
        await state.update_data(xodimlar_soni=xodimlar_soni, xodimlar_data=[], xodimlar=[])
        await ManualInput.xodimlar_data.set()
        await message.answer(translate_text(f"Xodimlar ma'lumotlarini kiriting (har bir xodim uchun: raqam (lavozim) – shu oy summasi so'm (yil boshidan jami so'm), masalan:\n1 (Rahbar) – 0 so'm (5000000 so'm)\nHar bir xodimni alohida kiriting, 1-xodimdan boshlang:"), lang)
    except ValueError:
        await message.answer(translate_text("❌ Xodimlar soni raqam bo'lishi kerak.", lang))

logger = logging.getLogger(__name__)

import re
import openpyxl
import logging

logger = logging.getLogger(__name__)

PHONE_RE = re.compile(r'^\+998\d{9}$')  # +998901234567 format

def _to_str(x):
    s = "" if x is None else str(x).strip()
    # Exceldagi 12.0 -> 12 bo'lib ketsin
    if s.endswith(".0"):
        try:
            s = str(int(float(s)))
        except:
            pass
    return s

def parse_firms_excel(file_path, lang='uz_latin'):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Agar sahifa nomi 'Лист1' bo'lsa — shu, aks holda active
        sheet = wb['Лист1'] if 'Лист1' in wb.sheetnames else wb.active

        firms = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            # ⚠️ 8 ta ustunni aynan shu tartibda o'qiymiz:
            # A      B            C           D        E        F        G        H
            # STIR,  Firma nomi,  Soliq turi, Rahbar, Telefon, DS%,     YS%,     QQS%
            try:
                stir, firma_nomi, soliq_turi, rahbar, phone, ds_stavka, ys_stavka, qqs_stavka = row
            except Exception as e:
                logger.warning(f"Ustunlar mos emas: {row}  ({e})")
                continue

            # ==== Validatsiyalar ====
            stir = _to_str(stir)
            firma_nomi = _to_str(firma_nomi)
            soliq_turi = _to_str(soliq_turi).lower()
            rahbar = _to_str(rahbar)
            phone = _to_str(phone)
            ds_stavka = _to_str(ds_stavka) or "0"
            ys_stavka = _to_str(ys_stavka) or "0"
            qqs_stavka = _to_str(qqs_stavka) or "0"

            if not (stir and firma_nomi and soliq_turi and rahbar and phone):
                logger.warning(f"Noto'g'ri qator (bo'sh maydon): {row}")
                continue

            if not re.match(r'^\d{9}$', stir):
                logger.warning(f"Noto'g'ri STIR: {stir}")
                continue

            if soliq_turi not in ("ds-ys", "ds-qqs"):
                logger.warning(f"Noto'g'ri soliq turi '{soliq_turi}', 'ds-ys' ga o'zgartirildi")
                soliq_turi = "ds-ys"

            if not PHONE_RE.match(phone):
                logger.warning(f"Noto'g'ri telefon raqami: {phone}")
                continue

            # Stavkalarni raqamga "ko'rinadigan" qilib normalize qilamiz (12, 4, 0 kabi)
            def norm_pct(s):
                s = s.replace("%", "")
                return s if s else "0"
            ds_stavka = norm_pct(ds_stavka)
            ys_stavka = norm_pct(ys_stavka)
            qqs_stavka = norm_pct(qqs_stavka)

            firms.append({
                "stir": stir,
                "firma_nomi": firma_nomi,
                "rahbar": rahbar,
                "soliq_turi": soliq_turi,
                "phone": phone,
                "ds_stavka": ds_stavka,
                "ys_stavka": ys_stavka,
                "qqs_stavka": qqs_stavka,
            })

        return firms, None

    except Exception as e:
        logger.error(f"Excel faylini o'qishda xato: {e}")
        return None, f"Excel faylini o'qishda xato: {str(e)}"

    


class AddFirmsFromExcel(StatesGroup):
    excel_upload = State()

@dp.callback_query_handler(lambda c: c.data == "add_firms_excel", user_id=ADMIN_IDS)
async def start_add_firms_excel(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    last_message_id = data.get('last_message_id')

    # Avvalgi xabarni o‘chirish
    if last_message_id:
        try:
            await bot.delete_message(chat_id=user_id, message_id=last_message_id)
        except Exception as e:
            logger.warning(f"Xabar o'chirishda xato: {e}, message_id={last_message_id}")

    await AddFirmsFromExcel.excel_upload.set()
    sent_message = await bot.send_message(
        callback_query.from_user.id,
        translate_text("xlsx faylni yuklang yoki /cancel bosib amaliyotni bekor qilin", lang)
    )
    await state.update_data(last_message_id=sent_message.message_id)



@dp.message_handler(content_types=['document'], state=AddFirmsFromExcel.excel_upload, user_id=ADMIN_IDS)
async def process_firms_excel(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    await state.update_data(user_id=user_id)  # user_id ni state ga saqlash
    if not message.document.file_name.endswith('.xlsx'):
        await message.answer(translate_text("❌ Faqat .xlsx fayllarni yuklang.", lang))
        await state.finish()
        return

    file_path = os.path.join(DATA_PATH, "temp", f"firms_{user_id}_{int(datetime.now().timestamp())}.xlsx")
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    await message.document.download(destination_file=file_path)

    firms, error = parse_firms_excel(file_path, lang)
    if not firms:
        await message.answer(translate_text(f"❌ Faylni o'qishda xato yuz berdi: {error}", lang))
        if os.path.exists(file_path):
            os.remove(file_path)
        await state.finish()
        return

    added_firms = []
    for firm in firms:
        stir = firm['stir']
        firma_nomi = firm['firma_nomi']
        rahbar = firm['rahbar']
        phone = firm['phone']
        soliq_turi = firm['soliq_turi'].lower()
        ds_stavka = firm['ds_stavka']
        ys_stavka = firm['ys_stavka']
        qqs_stavka = firm['qqs_stavka']

        add_firma(stir, firma_nomi, rahbar, soliq_turi, ds_stavka, ys_stavka, qqs_stavka)
        add_firm_owner(stir, phone)
        logger.info(f"Telefon saqlandi: {phone} -> STIR {stir}")
        try:
            os.makedirs(os.path.join(DATA_PATH, stir, "daromad"), exist_ok=True)
            if soliq_turi == 'ds-ys':
                os.makedirs(os.path.join(DATA_PATH, stir, "yagona"), exist_ok=True)
                logger.info(f"Papka yaratildi: data/{stir}/yagona")
            elif soliq_turi == 'ds-qqs':
                os.makedirs(os.path.join(DATA_PATH, stir, "qqs"), exist_ok=True)
                logger.info(f"Papka yaratildi: data/{stir}/qqs")
            else:
                logger.warning(f"Noto'g'ri soliq_turi: {soliq_turi} firma uchun {stir}")
        except Exception as e:
            logger.error(f"Papka yaratishda xato: {e}")
        added_firms.append(f"{firma_nomi} ({stir})")
        logger.info(f"Yangi firma qo'shildi: STIR={stir}, Name={firma_nomi}, Soliq_turi={soliq_turi}")

    if os.path.exists(file_path):
        os.remove(file_path)
        logger.info(f"Vaqtinchalik fayl o'chirildi: {file_path}")

    await message.answer(translate_text(f"✅ Quyidagi firmalar qo'shildi:\n{', '.join(added_firms)}", lang))
    await state.finish()
    await state.update_data(user_id=user_id)
    await back_to_admin_panel(state=state) # message orqali user_id uzatiladi

@dp.message_handler(state=ManualInput.xodimlar_data, user_id=ADMIN_IDS)
async def process_xodimlar_data(message: types.Message, state=FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    xodimlar_soni = data['xodimlar_soni']
    xodimlar_data = data.get('xodimlar_data', [])
    xodimlar = data.get('xodimlar', [])

    # Yangi format: tartib raqami, lavozim, ism familya, shu oy, yil boshidan
    pattern = r'^\s*(\d+)\s*\((.*?)\)\s*–\s*(.*?)\s*–\s*([\d\s]+)\s*so[''\']m\s*\(([\d\s]+)\s*so[''\']m\)\s*$'
    match = re.match(pattern, message.text.strip())

    if not match:
        await message.answer(
            translate_text(
                f"❌ Noto'g'ri format. Namuna: 1 (Rahbar) – Aliyev Valijon – 0 so'm (5000000 so'm)\n"
                f"Kiritilgan matn: {message.text.strip()}",
                lang
            )
        )
        logger.error(f"Noto'g'ri format kiritildi: user_id={user_id}, matn={message.text.strip()}")
        return

    index, lavozim, ism_familya, shu_oy, yil_boshidan = match.groups()
    index = int(index)
    shu_oy = int(shu_oy.replace(" ", ""))
    yil_boshidan = int(yil_boshidan.replace(" ", ""))

    if index != len(xodimlar_data) + 1:
        await message.answer(translate_text(f"❌ Noto'g'ri tartib raqami. {len(xodimlar_data) + 1}-xodimni kiriting.", lang))
        logger.error(f"Noto'g'ri tartib raqami: user_id={user_id}, index={index}, kutilgan={len(xodimlar_data) + 1}")
        return

    # Tilga qarab lavozim va ismni tarjima qilish
    if lang == 'uz_cyrillic':
        lavozim = convert_to_cyrillic(lavozim)
        ism_familya = convert_to_cyrillic(ism_familya)
    else:
        lavozim = translate_text(lavozim, 'uz_latin')
        ism_familya = translate_text(ism_familya, 'uz_latin')

    # To‘liq formatni shakllantirish
    xodimlar_data.append(
        f"{index} ({lavozim}) – {ism_familya}, "
        f"{translate_text('bu_oy_uchun_hisobotda', lang)}: {shu_oy:,} {translate_text('so‘m', lang)} "
        f"({translate_text('yil_boshidan_hisobotda', lang)}: {yil_boshidan:,} {translate_text('so‘m', lang)})"
    )
    xodimlar.append({
        'lavozim': lavozim,
        'ism': ism_familya,
        'yil_boshidan': yil_boshidan,
        'shu_oy': shu_oy
    })

    await state.update_data(xodimlar_data=xodimlar_data, xodimlar=xodimlar)
    logger.info(f"Xodim ma'lumotlari qo'shildi: user_id={user_id}, index={index}, lavozim={lavozim}, ism={ism_familya}")

    if len(xodimlar_data) < xodimlar_soni:
        await message.answer(translate_text(f"Keyingi xodim ma'lumotlarini kiriting ({len(xodimlar_data) + 1}/{xodimlar_soni}):", lang))
        return

    hisobot_davri_oylik = sum(x['shu_oy'] for x in xodimlar if x['shu_oy'] > 0)
    jami_oylik = sum(x['yil_boshidan'] for x in xodimlar)
    soliq = int(hisobot_davri_oylik * 0.12)
    firma_nomi = data['firma_name']
    oy = data['oy']

    result = get_text(lang, 'daromad_report',
                      firma_name=firma_nomi,
                      oy=get_month_name(lang, oy),
                      xodimlar_soni=xodimlar_soni,
                      xodimlar_data='\n'.join(xodimlar_data),
                      jami_oylik=jami_oylik,
                      hisobot_davri_oylik=hisobot_davri_oylik,
                      soliq=soliq)
    await state.update_data(
        xodimlar_soni=xodimlar_soni,
        hisobot_davri_oylik=hisobot_davri_oylik,
        jami_oylik=jami_oylik,
        soliq=soliq
    )
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton(translate_text("Tasdiqlash", lang), callback_data="confirm_report"),
        InlineKeyboardButton(translate_text("Tahrirlash", lang), callback_data="edit_report"),
        InlineKeyboardButton(translate_text("Bekor qilish", lang), callback_data="cancel_report")
    )
    await ManualInput.confirm.set()
    await message.answer(result + "\n" + translate_text("Tasdiqlaysizmi?", lang), reply_markup=keyboard)



@dp.callback_query_handler(lambda c: c.data == "confirm_report", user_id=ADMIN_IDS, state=ManualInput.confirm)
async def confirm_manual_report(callback_query: types.CallbackQuery, state=FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    stir = data.get('stir')
    oy = data.get('oy')
    soliq_turi = data.get('soliq_turi')

    # Retrieve firma_name from state or database
    firma_name = data.get('firma_name')
    if not firma_name:
        firma_name = get_firma_name(stir)
        if not firma_name:
            logger.error(f"Firma nomi topilmadi: STIR={stir}, user_id={user_id}")
            await bot.send_message(
                callback_query.from_user.id,
                translate_text("❌ Firma nomi topilmadi. Iltimos, qayta boshlang.", lang)
            )
            await state.finish()
            return
        await state.update_data(firma_name=firma_name)

    try:
        if soliq_turi == 'daromad':
            xodimlar_soni = data['xodimlar_soni']
            xodimlar_data = data['xodimlar_data']
            hisobot_davri_oylik = data['hisobot_davri_oylik']
            jami_oylik = data['jami_oylik']
            soliq = data['soliq']
            xodimlar = data.get('xodimlar', [])

            save_manual_report(stir, oy, firma_name, xodimlar_soni, "\n".join(xodimlar_data), hisobot_davri_oylik, jami_oylik, soliq)

            dest_path_latin = os.path.join(DATA_PATH, stir, "daromad", f"{get_month_name('uz_latin', oy)}1.xlsx")
            dest_path_cyrillic = os.path.join(DATA_PATH, stir, "daromad", f"{get_month_name('uz_cyrillic', oy)}1.xlsx")
            if generate_firma_excel(stir, oy, firma_name, xodimlar, dest_path_latin, dest_path_cyrillic):
                save_file(stir, "daromad", oy, "excel1_latin", dest_path_latin)
                save_file(stir, "daromad", oy, "excel1_cyrillic", dest_path_cyrillic)
                logger.info(f"Excel fayllari saqlandi: {dest_path_latin}, {dest_path_cyrillic}")
            else:
                logger.error(f"Excel fayllarini saqlashda xato: {dest_path_latin}, {dest_path_cyrillic}")
                await bot.send_message(
                    callback_query.from_user.id,
                    translate_text("⚠️ Hisobot saqlandi, lekin Excel fayllarini yaratishda xato yuz berdi.", lang)
                )
        elif soliq_turi == 'yagona':
            soliq_turi_yagona = data['soliq_turi_yagona']
            yil_boshidan_aylanma = data['yil_boshidan_aylanma']
            shu_oy_aylanma = data['shu_oy_aylanma']
            yagona_soliq = data['yagona_soliq']
            rahbar = data['rahbar']

            save_yagona_report(stir, oy, firma_name, rahbar, soliq_turi_yagona, yil_boshidan_aylanma, shu_oy_aylanma, yagona_soliq)

            dest_path_latin = os.path.join(DATA_PATH, stir, "yagona", f"{get_month_name('uz_latin', oy)}1.xlsx")
            dest_path_cyrillic = os.path.join(DATA_PATH, stir, "yagona", f"{get_month_name('uz_cyrillic', oy)}1.xlsx")
            if generate_yagona_excel(stir, oy, firma_name, rahbar, soliq_turi_yagona, yil_boshidan_aylanma, shu_oy_aylanma, dest_path_latin, dest_path_cyrillic):
                save_file(stir, "yagona", oy, "excel1_latin", dest_path_latin)
                save_file(stir, "yagona", oy, "excel1_cyrillic", dest_path_cyrillic)
                logger.info(f"Yagona Excel fayllari saqlandi: {dest_path_latin}, {dest_path_cyrillic}")
            else:
                logger.error(f"Yagona Excel fayllarini saqlashda xato: {dest_path_latin}, {dest_path_cyrillic}")
                await bot.send_message(
                    callback_query.from_user.id,
                    translate_text("⚠️ Hisobot saqlandi, lekin Excel fayllarini yaratishda xato yuz berdi.", lang)
                )
        elif soliq_turi == 'qqs':
            soliq_turi_qqs = data['soliq_turi_qqs']
            yil_boshidan_qqs = data['yil_boshidan_qqs']
            shu_oy_qqs = data['shu_oy_qqs']
            qqs_soliq = data['qqs_soliq']
            rahbar = data['rahbar']

            save_qqs_report(stir, oy, firma_name, rahbar, soliq_turi_qqs, yil_boshidan_qqs, shu_oy_qqs, qqs_soliq)

            dest_path_latin = os.path.join(DATA_PATH, stir, "qqs", f"{get_month_name('uz_latin', oy)}1.xlsx")
            dest_path_cyrillic = os.path.join(DATA_PATH, stir, "qqs", f"{get_month_name('uz_cyrillic', oy)}1.xlsx")
            if generate_yagona_excel(stir, oy, firma_name, rahbar, soliq_turi_qqs, yil_boshidan_qqs, shu_oy_qqs, dest_path_latin, dest_path_cyrillic):
                save_file(stir, "qqs", oy, "excel1_latin", dest_path_latin)
                save_file(stir, "qqs", oy, "excel1_cyrillic", dest_path_cyrillic)
                logger.info(f"QQS Excel fayllari saqlandi: {dest_path_latin}, {dest_path_cyrillic}")
            else:
                logger.error(f"QQS Excel fayllarini saqlashda xato: {dest_path_latin}, {dest_path_cyrillic}")
                await bot.send_message(
                    callback_query.from_user.id,
                    translate_text("⚠️ Hisobot saqlandi, lekin Excel fayllarini yaratishda xato yuz berdi.", lang)
                )
        else:
            logger.error(f"Noto'g'ri soliq_turi: {soliq_turi}, STIR={stir}, Oy={oy}")
            await bot.send_message(
                callback_query.from_user.id,
                translate_text("❌ Noto'g'ri soliq turi. Iltimos, qayta boshlang.", lang)
            )
            await state.finish()
            return

        excel_file_path = data.get('excel_file_path')
        if excel_file_path and os.path.exists(excel_file_path):
            os.remove(excel_file_path)
            logger.info(f"Vaqtinchalik fayl o'chirildi: {excel_file_path}")

        await state.finish()
        await bot.send_message(
            callback_query.from_user.id,
            translate_text(f"✅ {get_month_name(lang, oy)} uchun hisobot saqlandi.", lang)
        )
        logger.info(f"Hisobot saqlandi: STIR={stir}, Oy={oy}, Firma={firma_name}, Soliq_turi={soliq_turi}")
        await back_to_admin_panel(callback_query=callback_query, state=state)

    except Exception as e:
        logger.error(f"Hisobotni saqlashda xato: STIR={stir}, Oy={oy}, Soliq_turi={soliq_turi}, Error={str(e)}")
        await bot.send_message(
            callback_query.from_user.id,
            translate_text(f"❌ Hisobotni saqlashda xato yuz berdi: {str(e)}", lang)
        )
        await state.finish()

@dp.callback_query_handler(lambda c: c.data == "edit_report", user_id=ADMIN_IDS, state=ManualInput.confirm)
async def edit_manual_report(callback_query: types.CallbackQuery, state=FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    firma_name = data['firma_name']
    soliq_turi = data.get('soliq_turi')

    if soliq_turi == 'daromad':
        await ManualInput.firma_name.set()
        await bot.send_message(
            callback_query.from_user.id,
            translate_text(f"Firma nomi (hozirgi: {firma_name}, o'zgartirish uchun yangi nom kiriting yoki bo'sh qoldiring):", lang)
        )
    elif soliq_turi == 'yagona':
        await ManualInput.yagona_data.set()
        await bot.send_message(
            callback_query.from_user.id,
            translate_text(
                f"Yagona soliq ma'lumotlarini qayta kiriting (soliq stavkasi %, yil boshidan aylanma, shu oy aylanma, masalan: 4%, 10000000, 5000000):",
                lang
            )
        )
    elif soliq_turi == 'qqs':
        await ManualInput.qqs_data.set()
        await bot.send_message(
            callback_query.from_user.id,
            translate_text(
                f"QQS ma'lumotlarini qayta kiriting (soliq stavkasi %, yil boshidan QQS, shu oy QQS, masalan: 15%, 20000000, 10000000):",
                lang
            )
        )


@dp.callback_query_handler(lambda c: c.data == "cancel_report", user_id=ADMIN_IDS, state=ManualInput.confirm)
async def cancel_manual_report(callback_query: types.CallbackQuery, state=FSMContext):
    await bot.answer_callback_query(callback_query.id)
    user_id = callback_query.from_user.id
    lang = get_user_language(user_id)
    data = await state.get_data()
    excel_file_path = data.get('excel_file_path')
    if excel_file_path and os.path.exists(excel_file_path):
        os.remove(excel_file_path)
        logger.info(f"Vaqtinchalik fayl o'chirildi: {excel_file_path}")
    await state.finish()
    await bot.send_message(callback_query.from_user.id, translate_text("❌ Hisobot kiritish bekor qilindi.", lang))
    logger.info(f"Hisobot kiritish bekor qilindi: user_id={callback_query.from_user.id}")

    
@dp.message_handler(state=ManualInput.search, user_id=ADMIN_IDS)
async def process_search(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = get_user_language(user_id)
    search_query = message.text.strip().lower()

    data = await state.get_data()
    search_context = data.get("search_context")

    firms = get_all_firms()

    filtered_firms = []
    for stir, name in firms:
        if search_query in stir.lower() or search_query in name.lower():
            filtered_firms.append((stir, name))

    if not filtered_firms:
        await message.answer(translate_text("❌ Qidiruv bo‘yicha firma topilmadi.", lang))
        await state.finish()
        return

    keyboard, page, total_pages = create_paginated_keyboard(
        filtered_firms,
        search_context.replace("_search", ""),
        page=1,
        lang=lang
    )

    await bot.send_message(
        message.chat.id,
        translate_text(f"🔎 Qidiruv natijalari (Sahifa {page}/{total_pages}):", lang),
        reply_markup=keyboard
    )

    await state.finish()

